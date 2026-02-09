#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Fechamento Simples + Simulador walk-forward + Simulador de ciclos (ROI)

v11 (PATCH da v10):
- Novo modo --modo pool20:
- Novo modo --modo aposta16:
    - Gera 1 aposta de 16 dezenas (S16) a partir do mesmo POOL20 e padrões (resultado|moldura|metade|paridade).
    - Custo hardcoded: R$ 56,00 (aposta 16 dezenas na Lotofácil).

    - Gera 4 jogos de 15 a partir de um "POOL 20" (exclui 5 dezenas).
    - Nomes dos cartões no pool20: P1, P2, P3, P4 (em vez de AR/AS/BR/BS).
    - Default de --usar_cartoes no pool20: P1,P2,P3,P4
- Padrões para montar o POOL20 (excluir 5 dezenas):
    --pool20_padrao resultado|moldura|metade|paridade
    (default: resultado)
  Onde:
    - resultado: exclui 3 dentre as SORTEADAS do concurso base + 2 dentre as AUSENTES do base
    - moldura:   exclui 3 da MOLDURA + 2 do MIOLO
    - metade:    exclui 3 de 01..13 + 2 de 14..25
    - paridade:  exclui 3 ÍMPARES + 2 PARES
- Identificação clara dos arquivos gerados: inclui modo e padrão no nome do arquivo.
- Auto XLSX (sem precisar passar --resultados_xlsx) continua valendo:
    * usa resultados_DDMMYYYY.xlsx de hoje; senão ontem; senão baixa da CAIXA.
- Aba: por padrão tenta LOTOFÁCIL, e se não achar, usa a primeira aba.
  (Você ainda pode forçar com --aba, mas não precisa mais.)

Observação importante:
- 10 acertos NÃO paga. Todos os relatórios financeiros usam apenas >=11.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import unicodedata
import random
import glob
import statistics
import math
from dataclasses import dataclass
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


# =========================
# Normalização / utilitários
# =========================

UNIVERSO: Set[int] = set(range(1, 26))

# Custo da aposta de 16 dezenas (Lotofácil) — hardcoded por pedido.
APOSTA16_CUSTO: float = 56.00  # R$ 56,00
APOSTA17_CUSTO: float = 476.00  # R$ 476,00 (136 apostas de 15 a R$ 3,50)

def _strip_accents(s: str) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFD", s) if unicodedata.category(ch) != "Mn")

def _norm(s: str) -> str:
    s = _strip_accents((s or "").strip().lower())
    s = s.replace(" ", "")
    return s



def _fmt_nums(nums) -> str:
    """Formata uma coleção de dezenas (1..25) como '01 02 03 ...'"""
    if nums is None:
        return ""
    try:
        seq = sorted(int(x) for x in nums)
    except TypeError:
        # se vier um único int
        seq = [int(nums)]
    return " ".join(f"{n:02d}" for n in seq)

def _fmt_set(st) -> str:
    """Alias para _fmt_nums (mantido por compatibilidade)."""
    return _fmt_nums(st)

def _fmt_list(lst) -> str:
    """Formata lista de inteiros (ex.: excluídas) com zero à esquerda."""
    return _fmt_nums(lst)

def _sheet_match(wb, wanted: str) -> str:
    if wanted in wb.sheetnames:
        return wanted
    wanted_n = _norm(wanted)
    for name in wb.sheetnames:
        if _norm(name) == wanted_n:
            return name
    raise ValueError(f"Aba '{wanted}' não existe. Abas: {wb.sheetnames}")

def _pick_default_sheet(wb) -> str:
    # tenta LOTOFÁCIL primeiro
    for cand in ("LOTOFÁCIL", "LOTOFACIL"):
        for name in wb.sheetnames:
            if _norm(name) == _norm(cand):
                return name
    return wb.sheetnames[0]

def fmt_list(nums: Iterable[int]) -> str:
    return " ".join(f"{x:02d}" for x in sorted(nums))

def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def ddmmyyyy_today() -> str:
    return datetime.now().strftime("%d%m%Y")

def ddmmyyyy_yesterday() -> str:
    return (datetime.now() - timedelta(days=1)).strftime("%d%m%Y")

def to_mask(nums: Set[int]) -> int:
    m = 0
    for x in nums:
        if 1 <= x <= 25:
            m |= 1 << (x - 1)
    return m

def hits_mask(a_mask: int, b_mask: int) -> int:
    return (a_mask & b_mask).bit_count()

def _to_int(v) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        return None

def _to_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        try:
            return from_excel(v).date()
        except Exception:
            return None
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None

_money_re = re.compile(r"[-+]?\d[\d\.]*\,\d+|[-+]?\d[\d,]*\.\d+|[-+]?\d+")

def _to_float_money(v) -> Optional[float]:
    """
    Converte valores como:
      7
      7.00
      7,00
      R$ 7,00
      1.252,57
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip()
    if not s:
        return None
    m = _money_re.search(s.replace("R$", "").replace(" ", ""))
    if not m:
        return None
    num = m.group(0)

    if "," in num and "." in num:
        num = num.replace(".", "").replace(",", ".")
    else:
        if "," in num:
            num = num.replace(",", ".")
    try:
        return float(num)
    except ValueError:
        return None


# =========================
# Resultados (XLSX)
# =========================

@dataclass(frozen=True)
class Draw:
    concurso: int
    data: date
    bolas: Set[int]
    mask: int
    premios: Dict[int, float]  # {11:..., 12:..., 13:..., 14:..., 15:...}

    # Compat: versões antigas do script (e alguns modos) referenciam .date
    # enquanto este modelo usa .data. Mantemos alias para não quebrar outros modos.
    @property
    def date(self) -> date:
        return self.data

def _find_header_row(ws, max_scan_rows: int = 80) -> Tuple[int, Tuple]:
    """
    Procura cabeçalho nas primeiras N linhas.
    Exige: Concurso + Data + Bola1..Bola15
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        if not row:
            continue
        normed = [_norm(str(x) if x is not None else "") for x in row]
        has_concurso = "concurso" in normed
        has_data = ("datasorteio" in normed) or ("datadosorteio" in normed) or ("data" in normed)
        has_bolas = all(f"bola{i}" in normed for i in range(1, 16))
        if has_concurso and has_data and has_bolas:
            return row_idx, row
    raise ValueError("Não encontrei o cabeçalho (Concurso/Data Sorteio/Bola1..Bola15) nas primeiras linhas.")

def read_draws_xlsx(path: str, sheet_name: Optional[str] = None, diagnostico: bool = False) -> Tuple[List[Draw], str]:
    wb = load_workbook(path, data_only=True, read_only=False)

    if sheet_name:
        real_sheet = _sheet_match(wb, sheet_name)
        ws = wb[real_sheet]
    else:
        real_sheet = _pick_default_sheet(wb)
        ws = wb[real_sheet]

    header_row_idx, header = _find_header_row(ws)

    header_map: Dict[str, int] = {}
    for idx, col in enumerate(header):
        header_map[_norm(str(col) if col is not None else "")] = idx

    def col_idx(*names: str) -> int:
        for n in names:
            n2 = _norm(n)
            if n2 in header_map:
                return header_map[n2]
        raise ValueError(f"Coluna não encontrada no XLSX. Tentei: {names}")

    idx_concurso = col_idx("Concurso")
    idx_data = col_idx("Data Sorteio", "Data do Sorteio", "Data")
    idx_bolas = [col_idx(f"Bola{i}", f"Bola {i}") for i in range(1, 16)]

    # tenta achar colunas de prêmio (rateio/valor) para 11..15
    premio_cols: Dict[int, Optional[int]] = {11: None, 12: None, 13: None, 14: None, 15: None}
    norm_keys = list(header_map.keys())

    def find_premio_col(h: int) -> Optional[int]:
        for k in norm_keys:
            if ("rateio" in k or "valor" in k or "premio" in k) and str(h) in k:
                return header_map[k]
        return None

    for h in [11, 12, 13, 14, 15]:
        premio_cols[h] = find_premio_col(h)

    draws: List[Draw] = []
    total_linhas = 0
    linhas_ok = 0

    for r in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        total_linhas += 1
        if not r or len(r) <= max(idx_bolas):
            continue

        concurso = _to_int(r[idx_concurso])
        d = _to_date(r[idx_data])
        if concurso is None or d is None:
            continue

        bolas: List[int] = []
        ok = True
        for idx in idx_bolas:
            val = _to_int(r[idx])
            if val is None:
                ok = False
                break
            bolas.append(val)
        if not ok:
            continue

        if len(bolas) != 15 or len(set(bolas)) != 15 or any(x < 1 or x > 25 for x in bolas):
            continue

        premios: Dict[int, float] = {}
        for h in [11, 12, 13, 14, 15]:
            ci = premio_cols.get(h)
            v = _to_float_money(r[ci]) if (ci is not None and ci < len(r)) else None
            if v is not None:
                premios[h] = float(v)

        s = set(bolas)
        draws.append(Draw(concurso=concurso, data=d, bolas=s, mask=to_mask(s), premios=premios))
        linhas_ok += 1

    if diagnostico:
        print(f"[DIAG] Aba usada: {real_sheet}")
        print(f"[DIAG] Cabeçalho na linha: {header_row_idx}")
        print(f"[DIAG] Linhas lidas após cabeçalho: {total_linhas}")
        print(f"[DIAG] Concursos válidos carregados: {linhas_ok}")
        if draws and draws[-1].premios:
            print(f"[DIAG] Prêmios detectados (exemplo último): {draws[-1].premios}")
        else:
            print("[DIAG] Prêmios: não detectei colunas (payout ficará 0 se não houver).")

    if not draws:
        raise ValueError("Nenhum concurso encontrado no XLSX. (Aba/cabeçalho/dados não foram reconhecidos.)")

    draws.sort(key=lambda x: x.concurso)
    return draws, real_sheet


# =========================
# Download automático dos resultados
# =========================

CAIXA_URL = "https://servicebus3.caixa.gov.br/portaldeloterias/api/resultados/download?modalidade=Lotof%C3%A1cil"

def ensure_results_file(user_path: Optional[str]) -> str:
    """
    Se o usuário passou --resultados_xlsx, usa exatamente esse path.
    Senão:
      - procura resultados_DDMMYYYY.xlsx (hoje)
      - senão procura resultados_DDMMYYYY.xlsx (ontem)
      - senão baixa da CAIXA e salva como hoje
    """
    if user_path:
        return user_path

    today = ddmmyyyy_today()
    yesterday = ddmmyyyy_yesterday()
    f_today = f"resultados_{today}.xlsx"
    f_yest = f"resultados_{yesterday}.xlsx"

    if os.path.exists(f_today):
        print(f"[OK] Usando arquivo existente: {f_today}")
        return f_today
    if os.path.exists(f_yest):
        print(f"[OK] Usando arquivo do dia anterior: {f_yest}")
        return f_yest

    print("[DL] Baixando resultados mais recentes da CAIXA...")
    try:
        import requests
    except Exception as e:
        raise SystemExit("ERRO: para baixar automaticamente, instale requests: pip install requests") from e

    try:
        resp = requests.get(CAIXA_URL, timeout=45, verify=False)
    except Exception as e:
        raise SystemExit(f"ERRO: falha ao baixar arquivo da CAIXA: {e}") from e

    if resp.status_code != 200 or not resp.content:
        raise SystemExit(f"ERRO: download falhou (HTTP {resp.status_code}).")

    with open(f_today, "wb") as f:
        f.write(resp.content)

    print(f"[OK] Arquivo salvo como {f_today}")
    return f_today


# =========================
# ResultadoNorm (para facilitar search)
# =========================

def ensure_resultado_norm_column(xlsx_path: str, sheet_name: Optional[str], draws: List[Draw], sheet_used: str) -> None:
    """
    Adiciona/atualiza uma coluna 'ResultadoNorm' no XLSX (mesma linha do concurso),
    com a string "01 02 03 ... 25" (15 números), para você dar search/filtro fácil.

    Faz melhor esforço:
    - tenta localizar cabeçalho novamente e inserir a coluna ao final
    - não reordena nada; só escreve o texto normalizado
    """
    try:
        wb = load_workbook(xlsx_path, data_only=False, read_only=False)
        ws = wb[sheet_used]
        header_row_idx, header = _find_header_row(ws)
        # mapa colunas
        header_map = {}
        for idx, col in enumerate(header, start=1):
            header_map[_norm(str(col) if col is not None else "")] = idx
        # encontra colunas concurso e bola1
        idx_conc = header_map.get("concurso")
        if not idx_conc:
            return
        # coluna destino
        col_name = "ResultadoNorm"
        norm_key = _norm(col_name)
        dest_col = header_map.get(norm_key)
        if not dest_col:
            dest_col = ws.max_column + 1
            ws.cell(row=header_row_idx, column=dest_col, value=col_name)

        # monta lookup concurso->resultado
        lookup = {d.concurso: fmt_list(d.bolas) for d in draws}

        # percorre linhas e escreve
        updated = 0
        for r in range(header_row_idx + 1, ws.max_row + 1):
            conc = _to_int(ws.cell(row=r, column=idx_conc).value)
            if conc is None:
                continue
            val = lookup.get(conc)
            if not val:
                continue
            ws.cell(row=r, column=dest_col, value=val)
            updated += 1

        wb.save(xlsx_path)
        print(f"[OK] Coluna '{col_name}' atualizada: {updated} linhas em {xlsx_path}")
    except Exception as e:
        print(f"[WARN] Não consegui atualizar 'ResultadoNorm' no XLSX: {e}")


# =========================
# Ranking (freq/delay) para escolhas
# =========================

def _rank_frequency(draws: List[Draw], window: Optional[int] = None) -> Dict[int, int]:
    use = draws[-window:] if (window is not None and window > 0 and window < len(draws)) else draws
    freq = {i: 0 for i in range(1, 26)}
    for dr in use:
        for n in dr.bolas:
            freq[n] += 1
    return freq

def _rank_delay(draws: List[Draw], window: Optional[int] = None) -> Dict[int, int]:
    use = draws[-window:] if (window is not None and window > 0 and window < len(draws)) else draws
    last_seen = {i: None for i in range(1, 26)}
    for idx, dr in enumerate(use):
        for n in dr.bolas:
            last_seen[n] = idx
    max_idx = len(use) - 1
    delay = {}
    for n in range(1, 26):
        if last_seen[n] is None:
            delay[n] = len(use)
        else:
            delay[n] = max_idx - last_seen[n]
    return delay

def _score_number(n: int, freq: Dict[int, int], delay: Dict[int, int], mode: str) -> float:
    """
    Score para decidir o que manter/excluir.
    - mode=freq:       mais freq => maior score
    - mode=delay:      mais delay => maior score
    - mode=mixed:      (freq) + 0.25*(delay)
    """
    if mode == "freq":
        return float(freq.get(n, 0))
    if mode == "delay":
        return float(delay.get(n, 0))
    return float(freq.get(n, 0)) + 0.25 * float(delay.get(n, 0))

def _pick_exclusions(candidates: Iterable[int], k: int, freq: Dict[int, int], delay: Dict[int, int], rank_mode: str, seed: Optional[int]) -> List[int]:
    c = sorted(set(candidates))
    if k <= 0:
        return []
    if len(c) <= k:
        return c[:]
    # escolhe os "piores" scores para EXCLUIR (mantém os melhores)
    scored = [( _score_number(n, freq, delay, rank_mode), n) for n in c]
    scored.sort(key=lambda t: (t[0], t[1]))  # menores primeiro
    # desempate aleatório opcional: embaralha bloco final, mas sem perder determinismo
    if seed is not None:
        random.seed(seed)
    # pega k menores
    out = [n for _s, n in scored[:k]]
    return sorted(out)


# =========================
# Fechamento simples (AR/AS/BR/BS)
# =========================

@dataclass
class Closure:
    base_concurso: int
    base_data: date
    sorteadas: Set[int]
    nao_sorteadas: Set[int]
    fix_sorteadas: Set[int]
    fix_nao_sorteadas: Set[int]
    grupo_A: Set[int]
    grupo_B: Set[int]
    grupo_R: Set[int]
    grupo_S: Set[int]
    cartoes: Dict[str, Set[int]]  # AR, AS, BR, BS

def _split_remaining(remaining: List[int], size1: int, size2: int, seed: Optional[int], prefer_no_overlap: bool = True) -> Tuple[List[int], List[int]]:
    if seed is not None:
        random.seed(seed)
    rem = list(sorted(set(remaining)))
    random.shuffle(rem)
    a = rem[:size1]
    rem2 = [x for x in rem if x not in set(a)] if prefer_no_overlap else rem[:]
    b = rem2[:size2]
    if len(b) < size2:
        need = size2 - len(b)
        extra = [x for x in a if x not in set(b)]
        b.extend(extra[:need])
        b = b[:size2]
    return sorted(a), sorted(b)

def _choose_fixed_from_set(candidates: List[int], k: int, mode: str, draws_before: List[Draw], window: int, seed: Optional[int]) -> List[int]:
    if seed is not None:
        random.seed(seed)

    if k <= 0:
        return []

    cset = list(sorted(set(candidates)))
    if len(cset) < k:
        return cset[:]

    mode = (mode or "random").strip().lower()
    if mode == "random":
        return sorted(random.sample(cset, k))

    if mode in ("freq_window", "freq_all"):
        freq = _rank_frequency(draws_before, window if mode == "freq_window" else None)
        cset.sort(key=lambda n: (freq.get(n, 0), n), reverse=True)
        return sorted(cset[:k])

    if mode in ("delay_window", "delay_all"):
        delay = _rank_delay(draws_before, window if mode == "delay_window" else None)
        cset.sort(key=lambda n: (delay.get(n, 0), n), reverse=True)
        return sorted(cset[:k])

    return sorted(random.sample(cset, k))

def build_closure_for_base(base: Draw, draws_before_including_base: List[Draw], fix_s_mode: str, fix_n_mode: str, window: int, seed: Optional[int]) -> Closure:
    sorteadas = set(base.bolas)
    nao_sorteadas = set(range(1, 26)) - sorteadas

    fix_s = set(_choose_fixed_from_set(sorted(sorteadas), 3, fix_s_mode, draws_before_including_base, window, seed))
    fix_n = set(_choose_fixed_from_set(sorted(nao_sorteadas), 2, fix_n_mode, draws_before_including_base, window, seed))

    remaining_s = sorted(sorteadas - fix_s)
    a_extra, b_extra = _split_remaining(remaining_s, 6, 6, seed=seed, prefer_no_overlap=True)
    grupo_A = set(fix_s) | set(a_extra)
    grupo_B = set(fix_s) | set(b_extra)

    remaining_n = sorted(nao_sorteadas - fix_n)
    r_extra, s_extra = _split_remaining(remaining_n, 4, 4, seed=seed, prefer_no_overlap=True)
    grupo_R = set(fix_n) | set(r_extra)
    grupo_S = set(fix_n) | set(s_extra)

    cartoes = {
        "AR": set(grupo_A) | set(grupo_R),
        "AS": set(grupo_A) | set(grupo_S),
        "BR": set(grupo_B) | set(grupo_R),
        "BS": set(grupo_B) | set(grupo_S),
    }

    for nome, c in cartoes.items():
        if len(c) != 15:
            raise SystemExit(f"ERRO interno: cartão {nome} não tem 15 números (tem {len(c)}).")

    return Closure(
        base_concurso=base.concurso,
        base_data=base.data,
        sorteadas=sorteadas,
        nao_sorteadas=nao_sorteadas,
        fix_sorteadas=fix_s,
        fix_nao_sorteadas=fix_n,
        grupo_A=grupo_A,
        grupo_B=grupo_B,
        grupo_R=grupo_R,
        grupo_S=grupo_S,
        cartoes=cartoes,
    )


# =========================
# POOL20 (P1..P4)
# =========================

MOLDURA: Set[int] = {1,2,3,4,5,6,10,11,15,16,20,21,22,23,24,25}
MIOLO: Set[int] = UNIVERSO - MOLDURA
METADE_1_13: Set[int] = set(range(1,14))
METADE_14_25: Set[int] = set(range(14,26))
PARES: Set[int] = {2,4,6,8,10,12,14,16,18,20,22,24}
IMPARES: Set[int] = UNIVERSO - PARES

@dataclass
class Pool20Plan:
    padrao: str
    excluded: List[int]   # 5 excluídas
    pool20: List[int]     # 20 restantes (ordenadas)
    cartoes: Dict[str, Set[int]]  # P1..P4

def _pool20_make_games(pool20: List[int]) -> Dict[str, Set[int]]:
    """
    Constrói 4 jogos de 15 a partir de 20 dezenas.
    Método simples e estável:
      - divide as 20 dezenas em 5 blocos de 4: B0..B4
      - Jogo i (0..3): exclui o elemento i de cada bloco (5 exclusões) => sobra 15

    Isso produz 4 jogos (P1..P4) com exclusões distintas e bem distribuídas.
    """
    if len(pool20) != 20:
        raise SystemExit("ERRO interno: pool20 precisa ter 20 dezenas.")
    blocks = [pool20[i*4:(i+1)*4] for i in range(5)]
    games: Dict[str, Set[int]] = {}
    for i in range(4):
        excluded = {blocks[b][i] for b in range(5)}  # 5 números
        g = set(pool20) - excluded
        if len(g) != 15:
            raise SystemExit("ERRO interno: jogo pool20 não ficou com 15 números.")
        games[f"P{i+1}"] = g
    return games

def build_pool20_for_base(
    base: Draw,
    history_until_base: List[Draw],
    padrao: str,
    window: int,
    seed: Optional[int],
    rank_mode: str,
) -> Pool20Plan:
    """
    Monta POOL20 (20 dezenas) excluindo 5 conforme o padrão.
    Usa histórico (freq/delay) para decidir quais excluir dentro de cada subconjunto.
    """
    padrao = (padrao or "resultado").strip().lower()
    if padrao not in ("resultado", "moldura", "metade", "paridade"):
        raise SystemExit("ERRO: --pool20_padrao deve ser resultado|moldura|metade|paridade")

    freq = _rank_frequency(history_until_base, window if window and window > 0 else None)
    delay = _rank_delay(history_until_base, window if window and window > 0 else None)

    sorteadas = set(base.bolas)
    ausentes = UNIVERSO - sorteadas

    excluded: List[int] = []

    if padrao == "resultado":
        # exclui 3 dentre sorteadas + 2 dentre ausentes
        exc_s = _pick_exclusions(sorteadas, 3, freq, delay, rank_mode, seed)
        exc_a = _pick_exclusions(ausentes, 2, freq, delay, rank_mode, seed)
        excluded = sorted(exc_s + exc_a)

    elif padrao == "moldura":
        exc_m = _pick_exclusions(MOLDURA, 3, freq, delay, rank_mode, seed)
        exc_i = _pick_exclusions(MIOLO, 2, freq, delay, rank_mode, seed)
        excluded = sorted(exc_m + exc_i)

    elif padrao == "metade":
        exc_a = _pick_exclusions(METADE_1_13, 3, freq, delay, rank_mode, seed)
        exc_b = _pick_exclusions(METADE_14_25, 2, freq, delay, rank_mode, seed)
        excluded = sorted(exc_a + exc_b)

    elif padrao == "paridade":
        exc_imp = _pick_exclusions(IMPARES, 3, freq, delay, rank_mode, seed)
        exc_par = _pick_exclusions(PARES, 2, freq, delay, rank_mode, seed)
        excluded = sorted(exc_imp + exc_par)

    pool20_set = sorted(list(UNIVERSO - set(excluded)))
    if len(pool20_set) != 20:
        raise SystemExit(f"ERRO interno: pool20 ficou com {len(pool20_set)} dezenas (esperado 20).")

    games = _pool20_make_games(pool20_set)

    return Pool20Plan(padrao=padrao, excluded=excluded, pool20=pool20_set, cartoes=games)


# =========================
# Aposta 16 (S16) a partir do POOL20
# =========================

@dataclass
class Aposta16Plan:
    padrao: str
    excluded5: List[int]
    pool20: List[int]
    excluded4: List[int]
    s16: Set[int]
    cartoes: Dict[str, Set[int]]  # {"S16": set(...)}

def build_aposta16_for_base(
    base: Draw,
    history_until_base: List[Draw],
    padrao: str,
    window: int,
    seed: Optional[int],
    rank_mode: str,
) -> Aposta16Plan:
    """
    1) Monta POOL20 usando o mesmo pipeline do pool20 (exclui 5).
    2) A partir do POOL20, exclui mais 4 dezenas (ranking) => sobra S16 (16 dezenas).
    """
    plan20 = build_pool20_for_base(
        base=base,
        history_until_base=history_until_base,
        padrao=padrao,
        window=window,
        seed=seed,
        rank_mode=rank_mode,
    )

    freq = _rank_frequency(history_until_base, window if window and window > 0 else None)
    delay = _rank_delay(history_until_base, window if window and window > 0 else None)

    excluded4 = _pick_exclusions(plan20.pool20, 4, freq, delay, rank_mode, seed)
    s16 = set(plan20.pool20) - set(excluded4)
    if len(s16) != 16:
        raise SystemExit(f"ERRO interno: S16 ficou com {len(s16)} dezenas (esperado 16).")

    return Aposta16Plan(
        padrao=plan20.padrao,
        excluded5=plan20.excluded,
        pool20=plan20.pool20,
        excluded4=excluded4,
        s16=s16,
        cartoes={"S16": s16},
    )

def _aposta16_counts_from_k(k: int) -> Dict[int, int]:
    """
    Aposta 16 equivale a 16 jogos de 15 (cada um exclui 1 número da S16).
    Se k = |S16 ∩ sorteio15| então:
      - (16-k) jogos com k acertos  (exclui um número que NÃO saiu)
      - k jogos com (k-1) acertos  (exclui um número que SAIU)
    """
    k = int(k)
    if k < 0: k = 0
    if k > 15: k = 15
    counts: Dict[int, int] = {}
    counts[k] = counts.get(k, 0) + (16 - k)
    counts[k - 1] = counts.get(k - 1, 0) + k
    return counts


# =========================
# Premiação
# =========================


def payout_for_aposta16(draw: Draw, k_hits: int) -> float:
    """Payout total da aposta 16, somando os 16 jogos de 15 derivados, pagando só 11..15."""
    counts = _aposta16_counts_from_k(k_hits)
    total = 0.0
    for h, cnt in counts.items():
        if 11 <= int(h) <= 15 and int(cnt) > 0:
            total += float(draw.premios.get(int(h), 0.0)) * int(cnt)
    return float(total)

def payout_for_aposta17(draw, hits: int) -> float:
    """Approximate payout sum for a 17-number bet by expanding into C(17,15)=136 'jogos' of 15.
    Uses the same payout table as payout_for_hits() for each derived 15-number game.

    hits = |S17 ∩ resultado15|  (0..15)
    """
    # Aposta 17 picks 17 numbers; the draw has 15 numbers.
    # When you form each 15-number "jogo" by excluding 2 numbers from the 17,
    # you may exclude r of the hit numbers (0..2), so the derived jogo has hits-r.
    # Count of such jogos: C(hits, r) * C(17-hits, 2-r)
    from math import comb

    if hits < 0:
        return 0.0
    if hits > 15:
        hits = 15

    total = 0.0
    for r in (0, 1, 2):
        if r > hits:
            continue
        miss_in_set = 17 - hits
        if (2 - r) > miss_in_set:
            continue
        count = comb(hits, r) * comb(miss_in_set, 2 - r)
        total += count * payout_for_hits(draw, hits - r)
    return float(total)



def _percentile(values: List[float], p: float) -> float:
    """Simple percentile (0-100) with linear interpolation."""
    if not values:
        return 0.0
    xs = sorted(float(x) for x in values)
    if p <= 0:
        return xs[0]
    if p >= 100:
        return xs[-1]
    k = (len(xs) - 1) * (p / 100.0)
    f = math.floor(k)
    c = math.ceil(k)
    if f == c:
        return xs[int(k)]
    d0 = xs[f] * (c - k)
    d1 = xs[c] * (k - f)
    return d0 + d1



def _parse_percentiles(value, default: Tuple[float, float] = (40.0, 60.0)) -> Tuple[float, float]:
    """Parse percentile pair from CLI.

    Accepts:
      - "40,60" / "40 60" / "40;60" / "40|60"
      - sequences/tuples like (40, 60)
      - single number (uses it for both)

    Notes (pt-BR friendly):
      - Prefer treating comma as a *separator* when it clearly splits two tokens (e.g. "40,60").
      - If only one token is provided, comma may be used as decimal separator (e.g. "40,5").

    Returns (p_low, p_high) clamped to [0, 100] and ordered (low<=high).
    Falls back to *default* on any parsing error.
    """
    try:
        if value is None:
            p1, p2 = default

        elif isinstance(value, (tuple, list)):
            if len(value) >= 2:
                p1, p2 = float(value[0]), float(value[1])
            elif len(value) == 1:
                p1 = p2 = float(value[0])
            else:
                p1, p2 = default

        else:
            s = str(value).strip()
            if not s:
                p1, p2 = default
            else:
                # First try to split on common separators. If we get 2+ tokens, interpret as pair.
                parts = re.split(r"[\s;|]+", s)
                if len(parts) == 1 and "," in s:
                    # Comma may be separator for pair: "40,60"
                    comma_parts = [p.strip() for p in s.split(",") if p.strip() != ""]
                    if len(comma_parts) >= 2:
                        parts = comma_parts

                nums = []
                for part in parts:
                    if part:
                        nums.append(part)

                if len(nums) >= 2:
                    p1 = float(nums[0].replace(",", "."))
                    p2 = float(nums[1].replace(",", "."))
                elif len(nums) == 1:
                    # Single token: allow comma as decimal separator.
                    p1 = p2 = float(nums[0].replace(",", "."))
                else:
                    # Last resort: extract numbers (may treat comma as decimal separator).
                    found = re.findall(r"[-+]?\d+(?:[\.,]\d+)?", s)
                    if not found:
                        p1, p2 = default
                    elif len(found) == 1:
                        p1 = p2 = float(found[0].replace(",", "."))
                    else:
                        p1 = float(found[0].replace(",", "."))
                        p2 = float(found[1].replace(",", "."))

        # clamp + order
        p1 = max(0.0, min(100.0, float(p1)))
        p2 = max(0.0, min(100.0, float(p2)))
        if p1 > p2:
            p1, p2 = p2, p1
        return (p1, p2)

    except Exception:
        try:
            p1, p2 = float(default[0]), float(default[1])
            p1 = max(0.0, min(100.0, p1))
            p2 = max(0.0, min(100.0, p2))
            if p1 > p2:
                p1, p2 = p2, p1
            return (p1, p2)
        except Exception:
            return (40.0, 60.0)


def compute_aposta16_gate_stats(
    draws: List[Draw],
    *,
    pool20_padrao: str,
    pool20_rank: str,
    window: int,
    seed: Optional[int],
    min_hits: int = 12,
    teimosinha_n: int = 2,
    lookback_bases: int = 400,
    gap_percentis: Tuple[float, float] = (30.0, 70.0),
    metric: str = "concursos",
) -> Dict[str, object]:
    """Compute gating stats for Aposta16 strategy.

    The goal is to estimate:
      - win_rate: probability of hitting >=min_hits within the next teimosinha_n concursos
      - typical "gap" between *successful* events, and the current gap since the last success

    IMPORTANT: historically Lotofácil changed draw frequency (e.g., from conc. 2000 on it's ~6/wk),
    so using calendar days can distort gaps. Default metric is therefore **concursos**.
    metric:
      - "concursos": gaps measured in number of draws (recommended)
      - "dias": gaps measured in calendar days (legacy)

    Returns dict with p_low/p_high and current gap in the selected metric, plus an explanation string.
    """
    metric = (metric or "concursos").strip().lower()
    if metric not in ("concursos", "dias"):
        metric = "concursos"

    # Evaluate recent bases (walk-forward) and record when we get a success.
    n = len(draws)
    if n < 3:
        return {
            "ok": False,
            "metric": metric,
            "win_rate": 0.0,
            "wins": 0,
            "trials": 0,
            "p_low": 0.0,
            "p_high": 0.0,
            "gap_now": 0,
            "reason": "poucos concursos para calcular gate",
        }

    start = max(1, n - lookback_bases)  # base index starts at 1 (since base uses prev)
    wins = 0
    trials = 0
    success_targets: List[int] = []          # target indices where success happened (for concursos metric)
    success_target_dates: List[datetime] = []  # dates of those targets (for dias metric)

    for base_idx in range(start, n - 1):
        base = draws[base_idx - 1]
        plan16 = build_aposta16_for_base(
            base=base,
            history_until_base=draws[:max(0, base_idx - 1)],
            window=window,
            padrao=pool20_padrao,
            rank_mode=pool20_rank,
            seed=seed,
        )

        # search success in next teimosinha_n draws
        best_hits = -1
        best_when = None
        best_t_idx = None
        for off in range(1, teimosinha_n + 1):
            t_idx = base_idx + off
            if t_idx >= n:
                break
            target = draws[t_idx]
            k = len(set(plan16.s16) & set(target.bolas))
            if k > best_hits:
                best_hits = k
                best_when = target.data
                best_t_idx = t_idx

        trials += 1
        if best_hits >= min_hits:
            wins += 1
            # record the earliest target within teimosinha window that reaches min_hits
            # (we want gap distribution between actual hits)
            found = False
            for off in range(1, teimosinha_n + 1):
                t_idx = base_idx + off
                if t_idx >= n:
                    break
                target = draws[t_idx]
                k = len(set(plan16.s16) & set(target.bolas))
                if k >= min_hits:
                    success_targets.append(t_idx)
                    success_target_dates.append(target.data)
                    found = True
                    break
            if not found and best_t_idx is not None:
                # fallback: use best
                success_targets.append(best_t_idx)
                success_target_dates.append(best_when)

    win_rate = (wins / trials) if trials else 0.0

    # compute gaps between successive success events
    gaps: List[float] = []
    if metric == "concursos":
        success_targets_sorted = sorted(success_targets)
        for a, b in zip(success_targets_sorted, success_targets_sorted[1:]):
            gaps.append(float(b - a))
        gap_now = (n - 1) - success_targets_sorted[-1] if success_targets_sorted else (n - 1)
    else:
        dates_sorted = sorted(success_target_dates)
        for a, b in zip(dates_sorted, dates_sorted[1:]):
            gaps.append(float((b - a).days))
        gap_now = int((draws[-1].date - dates_sorted[-1]).days) if dates_sorted else int((draws[-1].date - draws[0].date).days)

    p_low = _percentile(gaps, float(gap_percentis[0])) if gaps else 0.0
    p_high = _percentile(gaps, float(gap_percentis[1])) if gaps else 0.0

    ok = (gap_now >= p_low) and (gap_now <= p_high)

    unit = "conc" if metric == "concursos" else "d"
    reason = (
        f"gap atual fora da faixa histórica ({p_low:.1f}-{p_high:.1f}{unit})"
        if not ok
        else "gap atual dentro da faixa histórica"
    )

    return {
        "ok": True,
        "metric": metric,
        "win_rate": win_rate,
        "wins": wins,
        "trials": trials,
        "successes": wins,
        "samples": trials,
        "p_low": p_low,
        "p_high": p_high,
        "gap_now": gap_now,
        "current_gap_days": gap_now if metric == "dias" else None,
        "current_gap_concursos": int(gap_now) if metric == "concursos" else None,
        "min_hits": min_hits,
        "teimosinha_n": teimosinha_n,
        "gate_pass": bool(ok),
        "gate_reason": reason,
    }
def payout_for_hits(draw: Draw, hits: int) -> float:
    if hits < 11 or hits > 15:
        return 0.0
    return float(draw.premios.get(hits, 0.0))


# =========================
# Simulação walk-forward (fechamento ou pool20)
# =========================

@dataclass
class CardStats:
    name: str
    played: bool
    hits_sum: int = 0
    n: int = 0
    ge11: int = 0
    ge12: int = 0
    ge13: int = 0
    ge14: int = 0
    ge15: int = 0
    payout_sum: float = 0.0

def _write_csv_dicts(path: str, rows: List[Dict[str, object]]) -> None:
    if not rows:
        with open(path, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(["info"])
            w.writerow(["(sem linhas)"])
        return

    fieldnames = list(rows[0].keys())
    s = set(fieldnames)
    for r in rows[1:]:
        for k in r.keys():
            if k not in s:
                fieldnames.append(k)
                s.add(k)

    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)

def _pct(n: int, d: int) -> float:
    return (100.0 * n / d) if d > 0 else 0.0

# =========================
# Engenharia reversa: gerar apostas sugeridas a partir do CSV de repetidos
# =========================

def _try_parse_date_br(s: str) -> Optional[date]:
    s = (s or '').strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, '%d/%m/%Y').date()
    except Exception:
        return None


def _parse_semicolon_ints(s: str) -> List[int]:
    s = (s or '').strip()
    if not s:
        return []
    out: List[int] = []
    for p in s.split(';'):
        p = str(p).strip()
        if not p:
            continue
        try:
            out.append(int(p))
        except ValueError:
            continue
    return out


def _parse_semicolon_strs(s: str) -> List[str]:
    s = (s or '').strip()
    if not s:
        return []
    return [p.strip() for p in s.split(';') if p.strip()]


def _find_latest_repeats_csv(prefix: str = 'simulacao_repetidos_', *, prefer_today: bool = True) -> Optional[str]:
    """Encontra automaticamente o CSV de repetidos mais recente.

    Regras:
      1) Se os arquivos seguirem o padrão do script (ex.: simulacao_repetidos_YYYYMMDD_HHMMSS.csv),
         escolhe pelo timestamp no nome do arquivo (mais confiável que mtime no Windows).
      2) Se prefer_today=True, tenta priorizar arquivos do dia de hoje (YYYYMMDD).
      3) Se não der para extrair timestamp do nome, cai para o mtime.
    """
    files = sorted(glob.glob(f'{prefix}*.csv'))
    if not files:
        return None

    today = datetime.now().strftime('%Y%m%d')

    # tenta usar timestamp do nome do arquivo
    # exemplo esperado: <prefix>20260118_015540.csv
    rx = re.compile(rf'^{re.escape(prefix)}(?P<d>\d{{8}})_(?P<t>\d{{6}})\.csv$', re.IGNORECASE)
    stamped: List[Tuple[str, str, str]] = []  # (YYYYMMDD, HHMMSS, path)
    for p in files:
        m = rx.match(os.path.basename(p))
        if not m:
            continue
        stamped.append((m.group('d'), m.group('t'), p))

    if stamped:
        # prioridade: arquivos de hoje
        if prefer_today:
            stamped_today = [x for x in stamped if x[0] == today]
            if stamped_today:
                stamped_today.sort(key=lambda x: (x[0], x[1]), reverse=True)
                return stamped_today[0][2]

        # senão, pega o mais novo global (pelo timestamp no nome)
        stamped.sort(key=lambda x: (x[0], x[1]), reverse=True)
        return stamped[0][2]

    # fallback: mtime
    files.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return files[0]


def generate_apostas_from_repetidos(
    repetidos_csv: str,
    *,
    top_n: int = 6,
    min_best_hits_today: int = 12,
    min_times_generated: int = 2,
    out_prefix: str = 'apostas',
) -> Tuple[str, List[Dict[str, object]]]:
    """
    Lê o CSV de repetidos (gerado pelo simulador), filtra candidatos que:
      - foram gerados pelo menos min_times_generated vezes
      - tiveram best_hits_today >= min_best_hits_today (ex.: 12+)
    E rankeia por:
      1) maior frequência (times_generated)
      2) menor min_gap entre aparições
      3) menor avg_gap
    Retorna path do relatório gerado + lista top.

    Observação:
    - Em aposta16 (S16), best_hits_today refere ao k=|S16 ∩ sorteio15| do dia.
    """
    rows: List[Dict[str, object]] = []

    with open(repetidos_csv, 'r', encoding='utf-8', newline='') as f:
        r = csv.DictReader(f)
        for row in r:
            try:
                times = int(float((row.get('times_generated') or '0').strip()))
            except Exception:
                times = 0
            try:
                best_hits = int(float((row.get('best_hits_today') or '0').strip()))
            except Exception:
                best_hits = 0
            try:
                sum_pay = float(str(row.get('sum_payout_today') or '0').replace(',', '.'))
            except Exception:
                sum_pay = 0.0

            if times < int(min_times_generated):
                continue
            if best_hits < int(min_best_hits_today):
                continue

            concs = _parse_semicolon_ints(row.get('all_target_concursos') or '')
            datas = _parse_semicolon_strs(row.get('all_target_datas') or '')
            concs = sorted(set(concs)) if concs else []

            gaps_conc: List[int] = []
            if len(concs) >= 2:
                for a, b in zip(concs, concs[1:]):
                    gaps_conc.append(int(b) - int(a))

            min_gap = min(gaps_conc) if gaps_conc else None
            avg_gap = (sum(gaps_conc) / len(gaps_conc)) if gaps_conc else None
            max_gap = max(gaps_conc) if gaps_conc else None

            # tenta gaps em dias (se tiver datas válidas)
            dates = [_try_parse_date_br(d) for d in datas]
            dates = [d for d in dates if d is not None]
            gaps_days: List[int] = []
            if len(dates) >= 2:
                dates_sorted = sorted(dates)
                for a, b in zip(dates_sorted, dates_sorted[1:]):
                    gaps_days.append((b - a).days)

            origin_day = (row.get('origin_target_data') or '').strip()

            rows.append({
                'card': (row.get('card') or '').strip(),
                'nums': (row.get('nums') or '').strip(),
                'times_generated': times,
                'origin_target_concurso': (row.get('origin_target_concurso') or '').strip(),
                'origin_target_data': origin_day,
                'best_hits_today': best_hits,
                'sum_payout_today': round(sum_pay, 2),
                'min_gap_concurso': min_gap if min_gap is not None else '',
                'avg_gap_concurso': round(avg_gap, 2) if avg_gap is not None else '',
                'max_gap_concurso': max_gap if max_gap is not None else '',
                'gaps_concurso': ';'.join(str(x) for x in gaps_conc),
                'min_gap_dias': min(gaps_days) if gaps_days else '',
                'avg_gap_dias': round(sum(gaps_days) / len(gaps_days), 2) if gaps_days else '',
                'max_gap_dias': max(gaps_days) if gaps_days else '',
                'gaps_dias': ';'.join(str(x) for x in gaps_days),
                'all_target_concursos': (row.get('all_target_concursos') or '').strip(),
                'all_target_datas': (row.get('all_target_datas') or '').strip(),
            })

    # rank: freq desc, min_gap asc, avg_gap asc, best_hits desc, payout desc
    def sort_key(x: Dict[str, object]):
        times = int(x.get('times_generated') or 0)
        min_gap = x.get('min_gap_concurso')
        avg_gap = x.get('avg_gap_concurso')
        best = int(x.get('best_hits_today') or 0)
        pay = float(x.get('sum_payout_today') or 0.0)

        # normaliza vazios
        mg = int(min_gap) if str(min_gap).strip() else 10**9
        ag = float(avg_gap) if str(avg_gap).strip() else 10**9
        return (-times, mg, ag, -best, -pay)

    rows_sorted = sorted(rows, key=sort_key)

    # regra: 1 jogo por dia (origin_target_data)
    picked: List[Dict[str, object]] = []
    seen_days: set[str] = set()
    for r in rows_sorted:
        day = str(r.get('origin_target_data') or '').strip()
        key = day if day else f"__nodate__{r.get('nums')}"
        if key in seen_days:
            continue
        seen_days.add(key)
        picked.append(r)
        if len(picked) >= int(top_n):
            break

    # escreve relatório
    stamp = now_stamp()
    out_csv = f"{out_prefix}_sugeridas_{stamp}.csv"

    # adiciona rank e escreve
    out_rows: List[Dict[str, object]] = []
    for i, r in enumerate(picked, start=1):
        rr = {'rank': i}
        rr.update(r)
        out_rows.append(rr)

    _write_csv_dicts(out_csv, out_rows)
    return out_csv, out_rows


# =========================
# Analise complementar (overlap 8/9/10 + estrategia A/B/C/D)
# =========================

def _parse_nums_str(nums: str) -> Set[int]:
    nums = (nums or '').strip()
    out: Set[int] = set()
    if not nums:
        return out
    for p in nums.replace(',', ' ').split():
        p = p.strip()
        if not p:
            continue
        try:
            out.add(int(p))
        except Exception:
            continue
    return out


def _calc_recent_freq_and_delay(draws: List[Draw], janela: int) -> Tuple[Dict[int, int], Dict[int, int]]:
    """Return (freq, delay).

    freq: count of each number in the last N contests (includes last contest).
    delay: how many contests since the last appearance (0 if it appeared in the last contest).
    """
    janela = int(janela or 0)
    window_draws = draws[-janela:] if janela and janela > 0 else draws

    freq: Dict[int, int] = {n: 0 for n in range(1, 26)}
    for d in window_draws:
        for n in d.bolas:
            freq[n] = freq.get(n, 0) + 1

    delay: Dict[int, int] = {n: 10**9 for n in range(1, 26)}
    for n in range(1, 26):
        dist = None
        for back, d in enumerate(reversed(draws)):
            if n in d.bolas:
                dist = back
                break
        delay[n] = int(dist) if dist is not None else 10**9

    return freq, delay


def _score_number_recent(n: int, freq: Dict[int, int], delay: Dict[int, int], janela: int) -> float:
    # Simple mixed score: recent frequency is the main signal; delay adds a small bias.
    f = float(freq.get(n, 0))
    d = float(delay.get(n, 0))
    j = float(janela if janela and janela > 0 else 50.0)
    d_norm = min(d, j) / j
    return (2.0 * f) + (1.0 * d_norm)


def _choose_group_A(last_result: Set[int], freq: Dict[int, int], delay: Dict[int, int], janela: int) -> Set[int]:
    # Choose 10 numbers from last result prioritizing recent frequency.
    ordered = sorted(list(last_result), key=lambda n: (-freq.get(n, 0), -_score_number_recent(n, freq, delay, janela), n))
    return set(ordered[:10])




def _choose_group_A_from_s16(
    last_result: Set[int],
    s16: Set[int],
    freq: Dict[int, int],
    delay: Dict[int, int],
    janela: int,
    fallback_A: Optional[Set[int]] = None,
) -> Set[int]:
    """Choose group A (10 dezenas) conditioned on the current S16.

    Strategy:
      - Start from intersection I = S16 ∩ last_result (the repeated numbers).
      - If |I| >= 10: pick the 10 best inside I by recent score.
      - Else: take all I and complete from (last_result \\ I) by recent score.
      - If something goes wrong / I empty, falls back to fallback_A or global _choose_group_A.
    """
    I = list(s16 & last_result)
    if not I:
        if fallback_A:
            return set(fallback_A)
        return _choose_group_A(last_result, freq, delay, janela)

    I_sorted = sorted(I, key=lambda n: (-_score_number_recent(n, freq, delay, janela), -freq.get(n, 0), n))
    picked = I_sorted[:10]
    if len(picked) < 10:
        rest = [n for n in last_result if n not in set(picked)]
        rest_sorted = sorted(rest, key=lambda n: (-_score_number_recent(n, freq, delay, janela), -freq.get(n, 0), n))
        for n in rest_sorted:
            if n not in picked:
                picked.append(n)
            if len(picked) >= 10:
                break

    # Safety: ensure size 10; if still not, fallback.
    if len(picked) != 10:
        if fallback_A:
            return set(fallback_A)
        return _choose_group_A(last_result, freq, delay, janela)

    return set(picked)
def _split_BCD(ordered_15: List[int]) -> Tuple[Set[int], Set[int], Set[int]]:
    # ordered_15 must have 15 items
    B = set(ordered_15[0:5])
    C = set(ordered_15[5:10])
    D = set(ordered_15[10:15])
    return B, C, D


def generate_overlap_analysis_for_top(
    top_rows: List[Dict[str, object]],
    draws: List[Draw],
    *,
    janela_recente: int = 40,
    A_por_s16: bool = False,
    out_prefix: str = 'simulacao',
) -> Tuple[str, List[Dict[str, object]]]:
    """Complementary analysis for suggested bets.

    For each suggested S16, compute:
      - overlap with the latest real contest (k_last)
      - historical overlap at the time it appeared (k_base_at = overlap with base contest on those dates)
      - suggest groups A/B/C/D and four 15-number games: A+B, A+C, A+D, B+C+D

    Returns (csv_path, rows).
    """
    if not top_rows:
        return '', []
    if len(draws) < 2:
        return '', []

    concurso_to_idx = {d.concurso: i for i, d in enumerate(draws)}
    last_draw = draws[-1]
    last_result = set(last_draw.bolas)

    freq, delay = _calc_recent_freq_and_delay(draws, int(janela_recente or 0))

    # Group A can be global (based on latest contest) OR conditioned per S16.
    # If A_por_s16 is True, we will derive A inside the loop for each S16 using S16 ∩ last_result.
    A_global = _choose_group_A(last_result, freq, delay, int(janela_recente or 0))

    def overlap_score(k: int) -> float:
        # prefer 9; 8 and 10 slightly lower; the rest progressively worse
        if k == 9:
            return 1.0
        if k in (8, 10):
            return 0.9
        if k in (7, 11):
            return 0.6
        return 0.2


    def overlap_points(k: int) -> float:
        # Aggressive weighting to enforce the 8/9/10 premise (peak at 9)
        if k == 9:
            return 1000.0
        if k in (8, 10):
            return 850.0
        if k in (7, 11):
            return 300.0
        if k in (6, 12):
            return 100.0
        return 0.0

    out: List[Dict[str, object]] = []

    for r in top_rows:
        s16 = _parse_nums_str(str(r.get('nums') or ''))
        k_last = len(s16 & last_result)

        concs = _parse_semicolon_ints(str(r.get('all_target_concursos') or ''))
        ks_at: List[int] = []
        for c in concs:
            idx = concurso_to_idx.get(int(c))
            if idx is None or idx <= 0:
                continue
            base = draws[idx - 1]
            ks_at.append(len(s16 & set(base.bolas)))

        pct_8_10 = _pct(sum(1 for x in ks_at if 8 <= x <= 10), len(ks_at))
        pct_eq9 = _pct(sum(1 for x in ks_at if x == 9), len(ks_at))
        avg_k_at = round(sum(ks_at) / len(ks_at), 2) if ks_at else ''
        best_k_at = max(ks_at) if ks_at else ''

        novas = sorted(list(s16 - last_result))
        novas_scored = sorted(novas, key=lambda n: (-_score_number_recent(n, freq, delay, int(janela_recente or 0)), -freq.get(n, 0), n))
        top6_novas = novas_scored[:6]
        avg_score_novas = round(sum(_score_number_recent(n, freq, delay, int(janela_recente or 0)) for n in novas) / len(novas), 4) if novas else 0.0

        # Choose A (10 dezenas)
        if A_por_s16:
            A = _choose_group_A_from_s16(last_result, s16, freq, delay, int(janela_recente or 0), fallback_A=A_global)
            A_mode = 'S16'
        else:
            A = set(A_global)
            A_mode = 'GLOBAL'

        complement_set = {n for n in range(1, 26) if n not in A}  # 15 dezenas
        complement_sorted = sorted(
            complement_set,
            key=lambda n: (-_score_number_recent(n, freq, delay, int(janela_recente or 0)), -freq.get(n, 0), n)
        )

        # Derive B/C/D per S16: prioritize novas from this S16, then complete from the complement of A
        B_list = []
        for n in novas_scored:
            if n in complement_set and n not in B_list:
                B_list.append(n)
            if len(B_list) >= 5:
                break
        if len(B_list) < 5:
            for n in complement_sorted:
                if n not in B_list:
                    B_list.append(n)
                if len(B_list) >= 5:
                    break
        B = set(B_list)
        remaining = [n for n in complement_sorted if n not in B]
        C = set(remaining[:5])
        D = set(remaining[5:10])

        jogo_AB = A | B
        jogo_AC = A | C
        jogo_AD = A | D
        jogo_BCD = B | C | D

        rr = dict(r)
        rr.update({
            'last_concurso': last_draw.concurso,
            'last_data': last_draw.data.strftime('%d/%m/%Y'),
            'k_last_overlap': k_last,
            'k_last_bucket': ('9' if k_last == 9 else ('8-10' if k_last in (8, 10) else 'outro')),
            'score_k_last': round(overlap_score(k_last), 3),
            'hist_count_aparicoes': len(ks_at),
            'hist_best_k_base': best_k_at,
            'hist_avg_k_base': avg_k_at,
            'hist_pct_k_8_10': round(pct_8_10, 2),
            'hist_pct_k_eq9': round(pct_eq9, 2),
            'novas_no_ultimo': ' '.join(f'{n:02d}' for n in novas),
            'novas_top6_score': ' '.join(f'{n:02d}' for n in top6_novas),
            'novas_avg_score': avg_score_novas,
            'grupo_A_10_ult': fmt_list(A),
            'grupo_A_modo': A_mode,
            'grupo_B_5': fmt_list(B),
            'grupo_C_5': fmt_list(C),
            'grupo_D_5': fmt_list(D),
            'jogo_A_B': fmt_list(jogo_AB),
            'jogo_A_C': fmt_list(jogo_AC),
            'jogo_A_D': fmt_list(jogo_AD),
            'jogo_B_C_D': fmt_list(jogo_BCD),
            'janela_recente': int(janela_recente or 0),
        })

        rr['score_final'] = round(
            overlap_points(k_last)
            + (float(rr.get('hist_pct_k_8_10') or 0.0) * 2.0)
            + (float(rr.get('hist_pct_k_eq9') or 0.0) * 1.0)
            + (avg_score_novas * 1.0)
            + (float(rr.get('times_generated') or 0) * 1.0),
            3
        )
        out.append(rr)

    out.sort(key=lambda x: (-float(x.get('score_final') or 0.0), -int(x.get('times_generated') or 0), -int(x.get('best_hits_today') or 0)))
    for i, rr in enumerate(out, start=1):
        rr['rank_overlap'] = i

    stamp = now_stamp()
    out_csv = f"{out_prefix}_sugeridas_overlap_{stamp}.csv"
    _write_csv_dicts(out_csv, out)
    return out_csv, out

def _calc_recent_overlap_stats(history: List[Draw], window: int = 40) -> Tuple[List[int], List[List[int]]]:
    """
    Calcula estatísticas simples de "overlap" entre concursos consecutivos dentro de uma janela.
    Retorna:
      - A_global: 5 dezenas mais recorrentes nos overlaps consecutivos (top-5)
      - A_groups: lista de grupos top-5 (global + grupos baseados em thresholds de overlap >= 8/9/10),
                 usado como candidatos para escolha do A (ex.: maximizar overlap com S16).
    Observação: função propositalmente conservadora (não cria novas features), serve apenas
    para suportar o modo ABCD sem quebrar outros modos.
    """
    if not history:
        return ([], [])

    # usa apenas a cauda para estabilidade
    h = history[-window:] if window and len(history) > window else history
    if len(h) < 2:
        # sem pares consecutivos
        # fallback: pega as dezenas do último concurso e completa com menores
        base = sorted(list(h[-1].bolas)) if h else []
        universo = list(range(1, 26))
        s = list(dict.fromkeys(base + universo))  # preserva ordem
        A = s[:5]
        return (A, [A])

    # contagem global por número em interseções consecutivas
    global_counts: Dict[int, int] = {n: 0 for n in range(1, 26)}
    pairs = []
    for i in range(1, len(h)):
        inter = set(h[i-1].bolas) & set(h[i].bolas)
        pairs.append(inter)
        for n in inter:
            global_counts[n] += 1

    def _top5_from_counts(counts: Dict[int, int]) -> List[int]:
        # maior contagem primeiro; desempate pelo número (menor primeiro)
        ranked = sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))
        top = [n for n, c in ranked if c > 0][:5]
        if len(top) < 5:
            # completa com menores que faltam (determinístico)
            for n in range(1, 26):
                if n not in top:
                    top.append(n)
                    if len(top) == 5:
                        break
        return top

    A_global = _top5_from_counts(global_counts)

    # grupos por thresholds de overlap >= 8/9/10 (espelha a ideia do comentário no script)
    groups: List[List[int]] = []
    for thr in (8, 9, 10):
        thr_counts: Dict[int, int] = {n: 0 for n in range(1, 26)}
        any_pair = False
        for inter in pairs:
            if len(inter) >= thr:
                any_pair = True
                for n in inter:
                    thr_counts[n] += 1
        if any_pair:
            g = _top5_from_counts(thr_counts)
            if g not in groups:
                groups.append(g)

    # garante que global esteja presente e na frente
    if A_global not in groups:
        groups.insert(0, A_global)
    else:
        # move para frente
        groups = [A_global] + [g for g in groups if g != A_global]

    return (A_global, groups)

def simulate_walk_forward(
    draws: List[Draw],
    modo: str,
    use_cards: List[str],
    custo_por_cartao: float,
    # fechamento
    fix_s_mode: str,
    fix_n_mode: str,
    # pool20
    pool20_padrao: str,
    pool20_rank_mode: str,
    # shared
    window: int,
    seed: Optional[int],
    out_prefix: str,
    repeats_min: int,
) -> Tuple[Dict[str, CardStats], float, float, str, Optional[str]]:
    """
    Walk-forward:
      Para cada concurso i (i>=1):
        - base = draws[i-1]
        - alvo = draws[i]
        - gera cartões com base no concurso anterior (base) + histórico até base
        - calcula acertos/payout no concurso alvo
    Gera CSV completo.
    """
    modo = (modo or "fechamento").strip().lower()
    if modo not in ("fechamento", "pool20", "aposta16"):
        raise SystemExit("ERRO: --modo deve ser fechamento|pool20|aposta16")

    all_cards = ["AR", "AS", "BR", "BS"] if modo == "fechamento" else (["P1", "P2", "P3", "P4"] if modo == "pool20" else ["S16"])
    use_cards_set = {c.upper().strip() for c in use_cards}
    for c in use_cards_set:
        if c not in all_cards:
            raise SystemExit(f"ERRO: cartão inválido em --usar_cartoes: {c}. Válidos: {','.join(all_cards)}")

    stats: Dict[str, CardStats] = {c: CardStats(name=c, played=(c in use_cards_set)) for c in all_cards}

    rows_full: List[Dict[str, object]] = []
    occurrences: Dict[Tuple[str, str], List[Dict[str, object]]] = {}
    payout_total = 0.0
    custo_total = 0.0

    for i in range(1, len(draws)):
        base = draws[i - 1]
        current = draws[i]
        history_until_base = draws[:i]  # inclui base

        if modo == "fechamento":
            closure = build_closure_for_base(
                base=base,
                draws_before_including_base=history_until_base,
                fix_s_mode=fix_s_mode,
                fix_n_mode=fix_n_mode,
                window=window,
                seed=seed,
            )
            cards = closure.cartoes
            meta = {
                "modo": "fechamento",
                "fixas_sorteadas_3": fmt_list(closure.fix_sorteadas),
                "fixas_nao_sorteadas_2": fmt_list(closure.fix_nao_sorteadas),
                "grupo_A": fmt_list(closure.grupo_A),
                "grupo_B": fmt_list(closure.grupo_B),
                "grupo_R": fmt_list(closure.grupo_R),
                "grupo_S": fmt_list(closure.grupo_S),
            }
        elif modo == "pool20":
            plan = build_pool20_for_base(
                base=base,
                history_until_base=history_until_base,
                padrao=pool20_padrao,
                window=window,
                seed=seed,
                rank_mode=pool20_rank_mode,
            )
            cards = plan.cartoes
            meta = {
                "modo": f"pool20:{plan.padrao}",
                "pool20_padrao": plan.padrao,
                "pool20_rank": pool20_rank_mode,
                "pool20_excluidas_5": fmt_list(plan.excluded),
                "pool20_20": fmt_list(plan.pool20),
            }
        else:
            plan16 = build_aposta16_for_base(
                base=base,
                history_until_base=history_until_base,
                padrao=pool20_padrao,
                window=window,
                seed=seed,
                rank_mode=pool20_rank_mode,
            )
            cards = plan16.cartoes
            meta = {
                "modo": f"aposta16:{plan16.padrao}",
                "pool20_padrao": plan16.padrao,
                "pool20_rank": pool20_rank_mode,
                "pool20_excluidas_5": fmt_list(plan16.excluded5),
                "pool20_20": fmt_list(plan16.pool20),
                "aposta16_excluidas_4": fmt_list(plan16.excluded4),
            }

        if modo == "aposta16":
            custo_concurso = (APOSTA16_CUSTO if ("S16" in use_cards_set) else 0.0)
        else:
            custo_concurso = len(use_cards_set) * float(custo_por_cartao)
        custo_total += custo_concurso

        total_payout_concurso = 0.0

        row = {
            "target_concurso": current.concurso,
            "target_data": current.data.strftime("%d/%m/%Y"),
            "base_concurso": base.concurso,
            "base_data": base.data.strftime("%d/%m/%Y"),
            "custo_concurso": round(custo_concurso, 2),
        }
        row.update(meta)

        for card in all_cards:
            nums = set(cards[card])
            hits = len(nums & current.bolas)
            if modo == "aposta16":
                payout = payout_for_aposta16(current, hits) if stats[card].played else 0.0
            else:
                payout = payout_for_hits(current, hits) if stats[card].played else 0.0

            cs = stats[card]
            cs.hits_sum += hits
            cs.n += 1
            if hits >= 11: cs.ge11 += 1
            if hits >= 12: cs.ge12 += 1
            if hits >= 13: cs.ge13 += 1
            if hits >= 14: cs.ge14 += 1
            if hits >= 15: cs.ge15 += 1
            if cs.played:
                cs.payout_sum += payout

            total_payout_concurso += payout

            row[f"{card}_nums"] = fmt_list(nums)
            row[f"{card}_hits"] = hits
            row[f"{card}_payout"] = round(payout, 2)
            if modo == "aposta16" and card == "S16":
                counts = _aposta16_counts_from_k(hits)
                for hh in range(11, 16):
                    row[f"S16_count_{hh}"] = int(counts.get(hh, 0))
            # store occurrence for repeats report (card + same 15 numbers)
            nums_str = fmt_list(nums)
            card_mask = to_mask(nums)
            key = (card, nums_str)
            occurrences.setdefault(key, []).append({
                "card": card,
                "nums": nums_str,
                "target_idx": i,
                "target_concurso": current.concurso,
                "target_data": current.data.strftime("%d/%m/%Y"),
                "base_concurso": base.concurso,
                "base_data": base.data.strftime("%d/%m/%Y"),
                "hits_today": hits,
                "payout_today": round(payout, 2),
                "mask": card_mask,
            })

        row["payout_total_concurso"] = round(total_payout_concurso, 2)
        row["lucro_liquido_concurso"] = round(total_payout_concurso - custo_concurso, 2)

        payout_total += total_payout_concurso
        rows_full.append(row)

    stamp = now_stamp()
    file_tag = modo if modo == "fechamento" else (f"pool20_{pool20_padrao}" if modo == "pool20" else f"aposta16_{pool20_padrao}")
    full_path = f"{out_prefix}_{file_tag}_{stamp}.csv"
    _write_csv_dicts(full_path, rows_full)

    
    # ===== Repetidos (mesmo cartão + mesmos 15 números gerados em múltiplos concursos) =====
    repeats_path: Optional[str] = None
    if repeats_min is None:
        repeats_min = 0
    if repeats_min < 1:
        repeats_min = 0

    if repeats_min >= 1:
        rows_repeats: List[Dict[str, object]] = []
        for (card, nums_str), occs in occurrences.items():
            if len(occs) < repeats_min:
                continue

            occs_sorted = sorted(occs, key=lambda x: int(x["target_idx"]))
            origin = occs_sorted[0]
            repeated = occs_sorted[1:]

            # métricas do "hoje" (nos concursos em que o jogo foi gerado)
            wins_ge11_today = sum(1 for o in occs_sorted if int(o["hits_today"]) >= 11)
            best_hits_today = max(int(o["hits_today"]) for o in occs_sorted) if occs_sorted else 0
            sum_payout_today = round(sum(float(o["payout_today"]) for o in occs_sorted), 2)

            # varredura futura a partir do 1º aparecimento (origem)
            origin_idx = int(origin["target_idx"])
            card_mask = int(origin["mask"])
            future_wins_ge11 = 0
            best_future_hits = 0
            sum_future_payout = 0.0
            future_15_concursos: List[int] = []
            future_15_datas: List[str] = []

            for j in range(origin_idx + 1, len(draws)):
                h = (card_mask & draws[j].mask).bit_count()
                if h >= 11:
                    future_wins_ge11 += 1
                    if h > best_future_hits:
                        best_future_hits = h
                    sum_future_payout += payout_for_hits(draws[j], h)
                    if h == 15:
                        future_15_concursos.append(draws[j].concurso)
                        future_15_datas.append(draws[j].data.strftime("%d/%m/%Y"))

            rowr: Dict[str, object] = {
                "card": card,
                "nums": nums_str,
                "times_generated": len(occs_sorted),
                "origin_target_concurso": origin["target_concurso"],
                "origin_target_data": origin["target_data"],
                "origin_base_concurso": origin["base_concurso"],
                "origin_base_data": origin["base_data"],
                "repeated_target_concursos": ";".join(str(o["target_concurso"]) for o in repeated),
                "repeated_target_datas": ";".join(str(o["target_data"]) for o in repeated),
                "repeated_base_concursos": ";".join(str(o["base_concurso"]) for o in repeated),
                "repeated_base_datas": ";".join(str(o["base_data"]) for o in repeated),
                "all_target_concursos": ";".join(str(o["target_concurso"]) for o in occs_sorted),
                "all_target_datas": ";".join(str(o["target_data"]) for o in occs_sorted),
                "wins_ge11_today": wins_ge11_today,
                "best_hits_today": best_hits_today,
                "sum_payout_today": sum_payout_today,
                "sum_future_wins_ge11": future_wins_ge11,
                "best_future_hits": best_future_hits,
                "sum_future_payout": round(sum_future_payout, 2),
                "future_15_concursos": ";".join(str(x) for x in future_15_concursos),
                "future_15_datas": ";".join(future_15_datas),
            }
            rows_repeats.append(rowr)

        if rows_repeats:
            repeats_path = f"{out_prefix}_repetidos_{stamp}.csv"
            _write_csv_dicts(repeats_path, rows_repeats)
    return stats, payout_total, custo_total, full_path, repeats_path

def print_sim_summary(stats: Dict[str, CardStats], payout_total: float, custo_total: float, use_cards: List[str]) -> None:
    used = [c.upper().strip() for c in use_cards]
    n_sim = stats[used[0]].n if used else 0

    print("\n=== SIMULAÇÃO (walk-forward) ===")
    print(f"Concursos simulados: {n_sim}  (pula o 1º — não tem concurso anterior)")
    print(f"Cartões jogados: {', '.join(used)}")
    print("\n=== RESUMO POR CARTÃO (>=11) ===")

    for card in used:
        cs = stats[card]
        mean_hits = (cs.hits_sum / cs.n) if cs.n else 0.0
        print(f"\n[{card}] média acertos: {mean_hits:.3f}")
        print(
            f"  >=11: {cs.ge11} ({_pct(cs.ge11, cs.n):.2f}%) | "
            f">=12: {cs.ge12} ({_pct(cs.ge12, cs.n):.2f}%) | "
            f">=13: {cs.ge13} ({_pct(cs.ge13, cs.n):.2f}%) | "
            f">=14: {cs.ge14} ({_pct(cs.ge14, cs.n):.2f}%) | "
            f">=15: {cs.ge15} ({_pct(cs.ge15, cs.n):.2f}%)"
        )
        print(f"  payout acumulado: R$ {cs.payout_sum:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

    net = payout_total - custo_total
    roi = (payout_total / custo_total) if custo_total > 0 else 0.0
    print("\n=== FINANCEIRO (somente cartões jogados) ===")
    print(f"Custo total:  R$ {custo_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    print(f"Payout total: R$ {payout_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    print(f"Líquido:      R$ {net:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    print(f"ROI bruto:    {roi:.3f}x")


# =========================
# Simulação de ciclos (estratégia: esperar após win, depois jogar janela)
# =========================

@dataclass
class CycleSummary:
    wait_after_win: int
    wait_after_loss: int
    trigger: str
    play_window: int
    teimosinha_n: int
    periodic: str
    cards: str
    played_contests: int
    wins_contests_ge11: int
    wins_ge11: int
    wins_ge12: int
    wins_ge13: int
    wins_ge14: int
    wins_ge15: int
    custo_total: float
    payout_total: float
    net_total: float
    roi: float
    avg_gap_between_wins: float
    max_gap_between_wins: int
    best_profit_streak: int
    best_loss_streak: int

def _wins_breakdown_from_hits(hits: int) -> Tuple[int,int,int,int,int]:
    return (
        1 if hits >= 11 else 0,
        1 if hits >= 12 else 0,
        1 if hits >= 13 else 0,
        1 if hits >= 14 else 0,
        1 if hits >= 15 else 0,
    )

def simulate_cycles_strategy(
    draws: List[Draw],
    *,
    modo: str,
    pool20_padrao: str,
    pool20_rank: str,
    wait_after_win: int,
    wait_after_loss: int,
    play_window: int,
    teimosinha_n: int,
    periodic: bool,
    use_cards: List[str],
    custo_por_cartao: float,
    fix_s_mode: str,
    fix_n_mode: str,
    window: int,
    seed: Optional[int],
    stop_on_win: bool = True,
) -> Tuple[CycleSummary, List[Dict[str, object]]]:
    if play_window <= 0:
        raise SystemExit("ERRO: play_window precisa ser >= 1 (use --ciclos_janelas).")

    modo = (modo or "fechamento").strip().lower()
    if modo not in ("fechamento", "pool20", "aposta16"):
        raise SystemExit("ERRO: --modo inválido para ciclos (use fechamento, pool20 ou aposta16).")

    all_cards = ["AR", "AS", "BR", "BS"] if modo == "fechamento" else (["P1", "P2", "P3", "P4"] if modo == "pool20" else ["S16"])
    use_cards_set = {c.upper().strip() for c in use_cards}
    for c in use_cards_set:
        if c not in all_cards:
            raise SystemExit(f"ERRO: cartão inválido: {c}. Válidos: {','.join(all_cards)}")

    trigger = "loss" if int(wait_after_loss) > 0 else ("periodic" if periodic else "win")

    played_contests = 0
    wins_contests_ge11 = 0
    wins_ge11 = wins_ge12 = wins_ge13 = wins_ge14 = wins_ge15 = 0
    payout_total = 0.0
    custo_total = 0.0

    last_win_i: Optional[int] = None
    gaps: List[int] = []

    best_profit_streak = 0
    best_loss_streak = 0
    cur_profit_streak = 0
    cur_loss_streak = 0

    # estado
    in_play = (int(wait_after_loss) <= 0)  # LOSS começa observando; WIN começa jogando
    play_left = (int(play_window) + int(teimosinha_n)) if in_play else 0
    cool_left = 0
    loss_streak = 0
    activate_play_next = False

    teimosinha_cache: Optional[Dict[str, Set[int]]] = None
    teimosinha_origin: Optional[Tuple[int, str]] = None  # (concurso, data)

    detalhe: List[Dict[str, object]] = []

    for i in range(1, len(draws)):
        base = draws[i - 1]
        current = draws[i]
        history_until_base = draws[:i]

        # gera cartões (para o concurso atual), mesmo em observação (modo LOSS precisa do "sinal")
        if modo == "fechamento":
            closure = build_closure_for_base(
                base=base,
                draws_before_including_base=history_until_base,
                fix_s_mode=fix_s_mode,
                fix_n_mode=fix_n_mode,
                window=window,
                seed=seed,
            )
            generated_cards = closure.cartoes
        elif modo == "pool20":
            plan = build_pool20_for_base(
                base=base,
                history_until_base=history_until_base,
                padrao=pool20_padrao,
                window=window,
                seed=seed,
                rank_mode=pool20_rank,
            )
            generated_cards = plan.cartoes
        else:
            plan16 = build_aposta16_for_base(
                base=base,
                history_until_base=history_until_base,
                padrao=pool20_padrao,
                window=window,
                seed=seed,
                rank_mode=pool20_rank,
            )
            generated_cards = plan16.cartoes

        # aplica teimosinha (repetir o mesmo jogo por N concursos dentro da janela)
        if in_play and int(teimosinha_n) > 0 and teimosinha_cache is not None:
            cards_today = teimosinha_cache
        else:
            cards_today = generated_cards
            if in_play and int(teimosinha_n) > 0 and teimosinha_cache is None:
                teimosinha_cache = cards_today
                teimosinha_origin = (current.concurso, current.data.strftime("%d/%m/%Y"))

        current_mask = current.mask

        per_card_hits: Dict[str, int] = {}
        any_ge11_virtual = False
        best_hits_virtual = 0

        for card in all_cards:
            nums = cards_today[card]
            h = hits_mask(to_mask(nums), current_mask)
            per_card_hits[card] = h
            best_hits_virtual = max(best_hits_virtual, h)
            if card in use_cards_set and h >= 11:
                any_ge11_virtual = True

        # ativa janela no próximo concurso
        if activate_play_next:
            in_play = True
            play_left = int(play_window) + int(teimosinha_n)
            activate_play_next = False
            teimosinha_cache = None
            teimosinha_origin = None

        # decide se joga hoje
        play_today = False
        if in_play:
            if cool_left > 0:
                cool_left -= 1
                play_today = False
            else:
                play_today = True

        payout_concurso = 0.0
        custo_concurso = 0.0
        net_concurso = 0.0
        win_today = False

        if play_today:
            played_contests += 1
            if modo == "aposta16":
                custo_concurso = (APOSTA16_CUSTO if ("S16" in use_cards_set) else 0.0)
            else:
                custo_concurso = float(custo_por_cartao) * len(use_cards_set)
            custo_total += custo_concurso

            for card in use_cards_set:
                h = per_card_hits[card]
                pay = payout_for_aposta16(current, h) if modo == "aposta16" else payout_for_hits(current, h)
                payout_concurso += pay
                if h >= 11:
                    win_today = True
                    wins_ge11 += 1
                    if h >= 12: wins_ge12 += 1
                    if h >= 13: wins_ge13 += 1
                    if h >= 14: wins_ge14 += 1
                    if h >= 15: wins_ge15 += 1

            payout_total += payout_concurso
            net_concurso = payout_concurso - custo_concurso

            if win_today:
                wins_contests_ge11 += 1
                if last_win_i is not None:
                    gaps.append(i - last_win_i)
                last_win_i = i

            if net_concurso > 0:
                cur_profit_streak += 1
                cur_loss_streak = 0
                best_profit_streak = max(best_profit_streak, cur_profit_streak)
            else:
                cur_loss_streak += 1
                cur_profit_streak = 0
                best_loss_streak = max(best_loss_streak, cur_loss_streak)

            play_left -= 1

            if win_today and stop_on_win:
                in_play = False
                play_left = 0
                teimosinha_cache = None
                teimosinha_origin = None
                cool_left = max(0, int(wait_after_win))
                loss_streak = 0

            if play_left <= 0 and in_play:
                in_play = False
                teimosinha_cache = None
                teimosinha_origin = None
                if periodic:
                    cool_left = max(0, int(wait_after_win))
                if int(wait_after_loss) > 0:
                    loss_streak = 0

        # modo LOSS: atualiza contagem em observação
        if int(wait_after_loss) > 0 and (not in_play) and (not play_today) and cool_left <= 0:
            if any_ge11_virtual:
                loss_streak = 0
            else:
                loss_streak += 1
            if loss_streak >= int(wait_after_loss):
                activate_play_next = True

        detalhe.append({
            "modo": modo,
            "concurso": current.concurso,
            "data": current.data.strftime("%d/%m/%Y"),
            "base_concurso": base.concurso,
            "base_data": base.data.strftime("%d/%m/%Y"),
            "trigger": trigger,
            "wait_after_win": int(wait_after_win),
            "wait_after_loss": int(wait_after_loss),
            "play_window": int(play_window),
            "teimosinha_n": int(teimosinha_n),
            "periodic": "SIM" if periodic else "NAO",
            "cards": ",".join(sorted(use_cards_set)),
            "played_today": "SIM" if play_today else "NAO",
            "best_hits_virtual": best_hits_virtual,
            "virtual_ge11_any": "SIM" if any_ge11_virtual else "NAO",
            "payout_concurso": round(payout_concurso, 2),
            "custo_concurso": round(custo_concurso, 2),
            "net_concurso": round(net_concurso, 2),
            "teimosinha_origin_concurso": teimosinha_origin[0] if teimosinha_origin else "",
            "teimosinha_origin_data": teimosinha_origin[1] if teimosinha_origin else "",
            **{f"{c}_hits": per_card_hits.get(c, "") for c in all_cards},
        })

    net_total = payout_total - custo_total
    roi = (payout_total / custo_total) if custo_total > 0 else 0.0
    avg_gap = (sum(gaps) / len(gaps)) if gaps else 0.0
    max_gap = max(gaps) if gaps else 0

    summary = CycleSummary(
        wait_after_win=int(wait_after_win),
        wait_after_loss=int(wait_after_loss),
        trigger=trigger,
        play_window=int(play_window),
        teimosinha_n=int(teimosinha_n),
        periodic="SIM" if periodic else "NAO",
        cards=",".join(sorted(use_cards_set)),
        played_contests=int(played_contests),
        wins_contests_ge11=int(wins_contests_ge11),
        wins_ge11=int(wins_ge11),
        wins_ge12=int(wins_ge12),
        wins_ge13=int(wins_ge13),
        wins_ge14=int(wins_ge14),
        wins_ge15=int(wins_ge15),
        custo_total=round(float(custo_total), 2),
        payout_total=round(float(payout_total), 2),
        net_total=round(float(net_total), 2),
        roi=round(float(roi), 4),
        avg_gap_between_wins=round(float(avg_gap), 2),
        max_gap_between_wins=int(max_gap),
        best_profit_streak=int(best_profit_streak),
        best_loss_streak=int(best_loss_streak),
    )
    return summary, detalhe

def print_last_preview(draws: List[Draw], modo: str, fix_s_mode: str, fix_n_mode: str, pool20_padrao: str, pool20_rank: str, window: int, seed: Optional[int]) -> None:
    if len(draws) < 2:
        raise SystemExit("ERRO: histórico insuficiente (precisa de pelo menos 2 concursos).")

    base = draws[-2]
    last = draws[-1]
    history_until_base = draws[:-1]

    print("\n=== PRÉVIA (último concurso como alvo) ===")
    print(f"Concurso base:   {base.concurso} | Data: {base.data.strftime('%d/%m/%Y')}")
    print(f"Concurso alvo:   {last.concurso} | Data: {last.data.strftime('%d/%m/%Y')}")
    print(f"Sorteadas base:      {fmt_list(base.bolas)}")
    print(f"Não sorteadas base:  {fmt_list(UNIVERSO - base.bolas)}")

    if modo == "fechamento":
        closure = build_closure_for_base(base, history_until_base, fix_s_mode, fix_n_mode, window, seed)

        print("\n=== Fixas escolhidas ===")
        print(f"Fixas SORTEADAS (3):     {fmt_list(closure.fix_sorteadas)}  (fix_s_mode={fix_s_mode}, window={window}, seed={seed})")
        print(f"Fixas NÃO SORTEADAS (2): {fmt_list(closure.fix_nao_sorteadas)}  (fix_n_mode={fix_n_mode}, window={window}, seed={seed})")
        print("\n=== Grupos do fechamento ===")
        print(f"A (9 sorteadas): {fmt_list(closure.grupo_A)}")
        print(f"B (9 sorteadas): {fmt_list(closure.grupo_B)}")
        print(f"R (6 não sort.): {fmt_list(closure.grupo_R)}")
        print(f"S (6 não sort.): {fmt_list(closure.grupo_S)}")

        print("\n=== Cartões finais ===")
        for nome in ["AR", "AS", "BR", "BS"]:
            nums = closure.cartoes[nome]
            repetidas = len(nums & base.bolas)
            nao = len(nums - base.bolas)
            print(f"{nome}: {fmt_list(nums)}  | verificação: repetidas={repetidas} não_sorteadas={nao}")

        print("\n=== Diagnóstico no concurso alvo ===")
        for nome in ["AR", "AS", "BR", "BS"]:
            nums = closure.cartoes[nome]
            h = len(nums & last.bolas)
            pay = payout_for_hits(last, h)
            if h >= 11:
                print(f"{nome}: {h} acertos | prêmio: R$ {pay:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            else:
                print(f"{nome}: {h} acertos")

    elif modo == "pool20":
        plan = build_pool20_for_base(base, history_until_base, pool20_padrao, window, seed, pool20_rank)
        print(f"\n=== POOL20 ({plan.padrao}) ===")
        print(f"Excluídas (5): {fmt_list(plan.excluded)}")
        print(f"POOL20 (20):   {fmt_list(plan.pool20)}")

        print("\n=== Jogos (P1..P4) ===")
        for nome in ["P1","P2","P3","P4"]:
            nums = plan.cartoes[nome]
            print(f"{nome}: {fmt_list(nums)}")

        print("\n=== Diagnóstico no concurso alvo ===")
        for nome in ["P1","P2","P3","P4"]:
            nums = plan.cartoes[nome]
            h = len(nums & last.bolas)
            pay = payout_for_hits(last, h)
            if h >= 11:
                print(f"{nome}: {h} acertos | prêmio: R$ {pay:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            else:
                print(f"{nome}: {h} acertos")

    else:
        plan16 = build_aposta16_for_base(base, history_until_base, pool20_padrao, window, seed, pool20_rank)
        print(f"\n=== APOSTA16 / S16 (a partir do POOL20: {plan16.padrao}) ===")
        print(f"POOL20 excluídas (5): {fmt_list(plan16.excluded5)}")
        print(f"POOL20 (20):          {fmt_list(plan16.pool20)}")
        print(f"S16 excluídas (4):    {fmt_list(plan16.excluded4)}")
        print(f"S16 (16):             {fmt_list(plan16.s16)}")

        print("\n=== Diagnóstico no concurso alvo ===")
        k = len(plan16.s16 & last.bolas)
        pay = payout_for_aposta16(last, k)
        counts = _aposta16_counts_from_k(k)
        if k >= 11:
            parts = []
            for hh in range(11, 16):
                c = counts.get(hh, 0)
                if c:
                    parts.append(f"{c}x{hh}")
            txt = ", ".join(parts) if parts else "-"
            print(f"S16: k={k} (acertos na S16) | combinação paga: {txt} | prêmio total: R$ {pay:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        else:
            print(f"S16: k={k} (acertos na S16)")



# =========================
# Main
# =========================

# =========================
# Simulador TOP6 diário + Gate (aposta16)
# =========================

def _score_s16_candidate(
    s16: Set[int],
    base_draw: Draw,
    history_until_base: List[Draw],
    window: int,
    rank_mode: str,
    overlap_target: int = 9,
    overlap_penalty: float = 35.0,
) -> Tuple[float, int]:
    """Score diário para rankear candidatos S16 usando só histórico até o base.
    - Soma score por número (freq/delay conforme rank_mode).
    - Penaliza desvio do overlap com o próprio concurso base (alvo típico=9).
    Retorna: (score, k_base)
    """
    freq = _rank_frequency(history_until_base, window if window and window > 0 else None)
    delay = _rank_delay(history_until_base, window if window and window > 0 else None)

    s = 0.0
    for n in sorted(s16):
        s += float(_score_number(n, freq, delay, rank_mode))
    k_base = len(set(s16) & set(base_draw.bolas))
    s -= float(overlap_penalty) * abs(int(k_base) - int(overlap_target))
    return float(s), int(k_base)


def _generate_top6_for_day(
    base_draw: Draw,
    history_until_base: List[Draw],
    padrao: str,
    window: int,
    seed0: int,
    rank_mode: str,
    candidates: int = 250,
    top_k: int = 6,
    overlap_target: int = 9,
    overlap_penalty: float = 0.0,
) -> List[Dict[str, Any]]:
    """Gera vários candidatos S16 (variando seed) e seleciona TOP-K únicos."""
    seen = set()
    scored_rows: List[Tuple[float, str, Dict[str, Any]]] = []

    for j in range(int(candidates)):
        sd = int(seed0) + int(base_draw.concurso) * 1000 + j
        plan16 = build_aposta16_for_base(
            base=base_draw,
            history_until_base=history_until_base,
            padrao=padrao,
            window=window,
            seed=sd,
            rank_mode=rank_mode,
        )
        s16 = set(plan16.s16)
        key = tuple(sorted(s16))
        if key in seen:
            continue
        seen.add(key)

        score, k_base = _score_s16_candidate(
            s16=s16,
            base_draw=base_draw,
            history_until_base=history_until_base,
            window=window,
            rank_mode=rank_mode,
            overlap_target=overlap_target,
            overlap_penalty=overlap_penalty,
        )
        row = {
            "card": "S16",
            "nums": fmt_list(s16),
            "s16_set": s16,
            "seed": sd,
            "k_base": k_base,
            "score": score,
            "padrao": padrao,
        }
        scored_rows.append((score, row["nums"], row))

    scored_rows.sort(key=lambda x: x[0], reverse=True)

    out: List[Dict[str, Any]] = []
    for _, _, r in scored_rows:
        out.append(r)
        if len(out) >= int(top_k):
            break
    return out





def compute_top6_gate_stats(
    *,
    draws: list,
    n_nums: int,
    padrao: str,
    rank_mode: str,
    window: int,
    top6_candidates: int,
    top6_size: int,
    teimosinha_n: int,
    min_hits: int,
    seed: int | None = None,
    overlap_penalty: float = 0.0,
    overlap_target: int = 9,
    gate_percentis=None,
    metric: str = "concursos",
) -> dict:
    """Walk-forward backtest for TOP6 S16/S17.

    For each base draw i, generates TOP6 candidate sets (size n_nums) using ONLY history up to i,
    then evaluates them on the next `teimosinha_n` contests (i+1..i+teimosinha_n).
    Marks success if any of the TOP6 achieves hits>=min_hits within that horizon.

    Returns:
      { 'rows': [...], 'summary_days': N, 'summary_success_days': K, 'summary_success_rate': K/N }
    """
    assert n_nums in (16, 17), "n_nums must be 16 or 17"
    teimosinha_n = max(1, int(teimosinha_n))
    top6_size = max(1, int(top6_size))
    top6_candidates = max(1, int(top6_candidates))

    def _parse_nums_str(s: str) -> set[int]:
        # '01 02 03 ...' -> {1,2,3,...}
        return {int(x) for x in str(s).strip().split() if x}

    def _payout(draw: "Draw", hits: int) -> float:
        if n_nums == 16:
            return float(payout_for_aposta16(draw, hits))
        if n_nums == 17:
            return float(payout_for_aposta17(draw, hits))
        return float(payout_for_hits(draw, hits))

    rows: list[dict] = []
    trials = 0
    successes = 0

    # Need at least 2 draws to evaluate (base + future).
    for i in range(len(draws) - 1):
        base = draws[i]
        history_until_base = draws[: i + 1]
        trials += 1

        # generate TOP6 for this base using history only
        top6_rows = _generate_top6_for_day_n(
            base_draw=base,
            history_until_base=history_until_base,
            padrao=padrao,
            rank_mode=rank_mode,
            window=window,
            candidates=top6_candidates,
            top6_size=top6_size,
            overlap_target=int(overlap_target),
            overlap_penalty=float(overlap_penalty),
            seed=seed,
            n_nums=n_nums,
        )

        # Evaluate on next contests (teimosinha horizon)
        best_hit = -1
        best_payout = 0.0
        best_when = None
        best_rank = None
        best_nums = None
        k_base_best = None

        hit_ge_min = False
        hit_ge_min_when = None

        first_ge_min_offset = None
        first_ge_min_payout = 0.0
        first_ge_min_hit = -1
        first_ge_min_when = None

        # future idxs
        max_j = min(len(draws) - 1, i + teimosinha_n)
        for j in range(i + 1, max_j + 1):
            target = draws[j]
            target_set = set(target.bolas)

            # melhor payout/hit entre os TOP6 neste concurso j
            best_pay_this = 0.0
            best_hit_this = -1
            for rnk, rr in enumerate(top6_rows, start=1):
                nums_set = _parse_nums_str(rr.get("nums", ""))
                k = len(nums_set & target_set)
                pay = _payout(target, k)

                if pay > best_pay_this or (pay == best_pay_this and k > best_hit_this):
                    best_pay_this = pay
                    best_hit_this = k

                if k > best_hit or (k == best_hit and pay > best_payout):
                    best_hit = k
                    best_payout = pay
                    best_when = target.data
                    best_rank = rnk
                    best_nums = rr.get("nums", "")
                    k_base_best = rr.get("k_base", 0)

                # first >=min_hits (para custo teimosinha)
                if (first_ge_min_offset is None) and (k >= min_hits):
                    first_ge_min_offset = int(j - i)
                    first_ge_min_when = target.data
                    first_ge_min_hit = k
                    first_ge_min_payout = float(best_pay_this)

            if (first_ge_min_offset is not None) and (not hit_ge_min):
                hit_ge_min = True
                hit_ge_min_when = first_ge_min_when
        if hit_ge_min:
            successes += 1

        # record row (base -> first target is i+1)
        target0 = draws[i + 1]
        rows.append(
            {
                "base_concurso": base.concurso,
                "base_index": i,
                "base_data": base.data,
                "n_nums": n_nums,
                "alvo_concurso": target0.concurso,
                "alvo_data": target0.data,
                "top6_size": top6_size,
                "top6_candidates": top6_candidates,
                "teimosinha_n": teimosinha_n,
                "min_hits": min_hits,
                "best_hit": best_hit,
                "best_payout": round(float(best_payout), 2),
                "best_when": best_when,
                "best_rank_in_top6": best_rank,
                "best_nums": best_nums,
                "k_base_best": k_base_best,
                "success_ge_min": 1 if hit_ge_min else 0,
                "first_ge_min_offset": first_ge_min_offset,
                "first_ge_min_payout": round(float(first_ge_min_payout), 2),
                "first_ge_min_when": first_ge_min_when,
            }
        )

    rate = (successes / trials) if trials else 0.0
    return {
        "rows": rows,
        "summary_days": trials,
        "summary_success_days": successes,
        "summary_success_rate": rate,
    }


def simulate_top6_gate(
    draws: List[Draw],
    n_nums: int,
    padrao: str,
    rank_mode: str,
    window: int,
    seed: int,
    top6_candidates: int,
    top6_size: int,
    teimosinha_n: int,
    min_hits: int,
    overlap_target: int,
    overlap_penalty: float,
    gate_percentis: Tuple[float, float],
    gate_min_trials: int,
    out_prefix: str = "",
    metric: str = "concursos",
) -> Tuple[str, Dict[str, Any]]:
    """Simulação walk-forward TOP6 + Gate (S16/S17), no mesmo modelo do ABCD:
    - calcula PASS por dia (gate dinâmico usando gaps em concursos)
    - aplica custo/payout/profit somente nos dias PASS
    - exporta CSV e devolve summary completo
    """

    # Para S16/S17 (uma aposta por concurso), simulamos somente TOP1.
    if int(n_nums) in (16, 17):
        top6_size = 1

    def _parse_date_any(v):
        if v is None:
            return None
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        if isinstance(v, str):
            # aceita dd/mm/yyyy e yyyy-mm-dd
            vv = v.strip()
            for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                try:
                    return datetime.strptime(vv, fmt).date()
                except Exception:
                    pass
        return None

    stats = compute_top6_gate_stats(
        draws=draws,
        n_nums=int(n_nums),
        padrao=padrao,
        rank_mode=rank_mode,
        top6_candidates=int(top6_candidates),
        top6_size=int(top6_size),
        teimosinha_n=int(teimosinha_n),
        min_hits=int(min_hits),
        seed=int(seed),
        metric=metric,
        window=int(window),
        overlap_penalty=float(overlap_penalty),
        overlap_target=int(overlap_target),
    )
    rows = stats["rows"]

    aposta_custo = None  # custo é inferido por concurso a partir do custo da aposta de 15

    # --- Aplica gate dinâmico por dia e calcula custo/payout/profit ---
    win_positions_all: List[int] = []  # índices (na lista rows) onde houve lucro>0 se tivesse jogado (independente do gate)
    rows_out: List[Dict[str, Any]] = []

    dias_pass = 0
    dias_skip = 0
    custo_total_pass = 0.0
    payout_total_pass = 0.0
    profit_total_pass = 0.0

    custo_total_all = 0.0
    payout_total_all = 0.0
    profit_total_all = 0.0

    for idx, r in enumerate(rows):
        # --- 1) Simula custo/payout/profit SEM depender do gate (modelo ABCD-like) ---
        base_i = int(r.get("base_index", 0))
        nums_set = _parse_nums_str(r.get("best_nums", ""))  # conjunto da aposta escolhida (TOP1)

        # Se atingiu min_hits em algum ponto, pode parar teimosinha no primeiro sucesso (quando houver offset válido)
        stop_at = None
        if int(r.get("success_ge_min", 0)) == 1:
            off = r.get("first_ge_min_offset")
            if isinstance(off, int) and off > 0:
                stop_at = base_i + off

        max_j = base_i + int(teimosinha_n)
        if stop_at is not None:
            max_j = min(max_j, int(stop_at))
        max_j = min(max_j, len(draws) - 1)

        custo = 0.0
        payout = 0.0
        contests_played = 0
        rep_details = []

        for j in range(base_i + 1, max_j + 1):
            d = draws[j]
            hits = len(nums_set & set(d.bolas))
            contests_played += 1

            payout_inc = 0.0
            custo_inc = 0.0

            if int(n_nums) == 16:
                payout_inc = payout_for_aposta16(d, hits)
                custo_inc = _infer_aposta16_custo(d, fallback15=APOSTA15_CUSTO_DEFAULT)
            elif int(n_nums) == 17:
                payout_inc = payout_for_aposta17(d, hits)
                custo_inc = _infer_aposta17_custo(d, fallback15=APOSTA15_CUSTO_DEFAULT)
            else:
                payout_inc = payout_for_hits(d, hits)
                custo_inc = _infer_aposta15_custo(d, fallback=APOSTA15_CUSTO_DEFAULT)

            payout += payout_inc
            custo += custo_inc

            rep_details.append({
                "j": j,
                "concurso": getattr(d, "concurso", None),
                "data": getattr(d, "data", None),
                "hits": int(hits),
                "custo": float(custo_inc),
                "payout": float(payout_inc),
                "profit": float(payout_inc - custo_inc),
            })

        profit = payout - custo
        win_if_played = 1 if (profit > 0.0) else 0

        # --- 2) Gate dinâmico baseado em gaps de WIN (lucro>0) calculados SEM depender do gate ---
        gate_enabled = True
        gate_pass = True
        gate_lo = 0.0
        gate_hi = 0.0
        gap_atual = 0

        if len(win_positions_all) < int(gate_min_trials):
            gate_enabled = False
            gate_pass = True
        else:
            gaps = [int(b - a) for a, b in zip(win_positions_all, win_positions_all[1:])]
            gap_atual = int(idx - win_positions_all[-1])
            p_low, p_high = float(gate_percentis[0]), float(gate_percentis[1])
            gate_lo = float(_percentile(gaps, p_low)) if gaps else 0.0
            gate_hi = float(_percentile(gaps, p_high)) if gaps else 0.0
            gate_pass = (gap_atual >= gate_lo and gap_atual <= gate_hi)

        if win_if_played == 1:
            win_positions_all.append(idx)

        played = bool(gate_pass)

        if played:
            dias_pass += 1
            custo_total_pass += custo
            payout_total_pass += payout
            profit_total_pass += profit
        else:
            dias_skip += 1

        # Totais "all" = hipotético se jogasse todo dia (igual ABCD: calcula independente do gate)
        custo_total_all += custo
        payout_total_all += payout
        profit_total_all += profit

        rr = dict(r)
        rr.update(
            {
                "gate_pass": bool(gate_pass),
                "played": 1 if played else 0,
                "contests_played": contests_played,
                "rep_detail": "|".join(
                    f"{k+1}:{rd.get('hits')}/{rd.get('payout'):.2f}/{rd.get('custo'):.2f}"
                    for k, rd in enumerate(rep_details)
                ),
                "custo": round(custo, 2),
                "payout": round(payout, 2),
                "profit": round(profit, 2),
                "win_profit_gt0": win_if_played,
                "gate_gap_atual": gap_atual,
                "gate_lo": round(gate_lo, 2),
                "gate_hi": round(gate_hi, 2),
            }
        )
        rows_out.append(rr)

    dias_avaliados = len(rows_out)
    sucessos = sum(1 for rr in rows_out if int(rr.get("win_profit_gt0", 0)) == 1)
    taxa = (sucessos / dias_avaliados) if dias_avaliados else 0.0
    winrate_pass = (sucessos / dias_pass) if dias_pass else 0.0

    profit_medio = (profit_total_all / dias_avaliados) if dias_avaliados else 0.0
    profit_medio_pass = (profit_total_pass / dias_pass) if dias_pass else 0.0

    # gate "atual" (último dia)
    gate_now = {
        "metric": "concursos",
        "percentis": (float(gate_percentis[0]), float(gate_percentis[1])),
        "faixa": (rows_out[-1].get("gate_lo", 0.0), rows_out[-1].get("gate_hi", 0.0)) if rows_out else (0.0, 0.0),
        "gap_atual": rows_out[-1].get("gate_gap_atual", 0) if rows_out else 0,
        "pass": bool(rows_out[-1].get("gate_pass", True)) if rows_out else True,
    }

    # export CSV
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"{out_prefix}_top6_gate_S{n_nums}_{ts}.csv" if out_prefix else f"top6_gate_S{n_nums}_{ts}.csv"
    csv_path = os.path.join(os.getcwd(), fname)
    _write_csv_dicts(csv_path, rows_out)

    summary = {
        "days": dias_avaliados,
        "success_days": sucessos,
        "success_rate": taxa,
        "dias_pass": dias_pass,
        "dias_skip": dias_skip,
        "winrate_pass": winrate_pass,
        "custo_total_pass": round(custo_total_pass, 2),
        "payout_total_pass": round(payout_total_pass, 2),
        "profit_total_pass": round(profit_total_pass, 2),
        "custo_total_all": round(custo_total_all, 2),
        "payout_total_all": round(payout_total_all, 2),
        "profit_total_all": round(profit_total_all, 2),
        "profit_medio": round(profit_medio, 2),
        "profit_medio_pass": round(profit_medio_pass, 2),
        "gate": gate_now,
        "gate_enabled": True,
        "gate_metric": "concursos",
        "gate_percentis": f"{float(gate_percentis[0])},{float(gate_percentis[1])}",
        "gate_p_low_conc": gate_now["faixa"][0],
        "gate_p_high_conc": gate_now["faixa"][1],
        "gap_now_conc": gate_now["gap_atual"],
        # campos legacy para não quebrar prints antigos
        "gate_p_low_days": 0.0,
        "gate_p_high_days": 0.0,
        "gap_now_days": 0,
        "gate_pass": gate_now["pass"],
        "gate_reason": "",
    }
    return csv_path, summary
def build_aposta17_for_base(
    base: Draw,
    history_until_base: list,
    padrao: str,
    rank_mode: str,
    window: int = 40,
    seed: int = 1,
) -> Aposta17Plan:
    plan20 = build_pool20_for_base(
        base=base,
        history_until_base=history_until_base,
        padrao=padrao,
        rank_mode=rank_mode,
        window=window,
        seed=seed,
    )
    # _rank_frequency/_rank_delay nesta base recebem apenas a lista de concursos
    # (history_until_base) e calculam rankings para 1..25.
    freq_rank = _rank_frequency(history_until_base, window=window)
    delay_rank = _rank_delay(history_until_base, window=window)
    excluded3 = _pick_exclusions(
        candidates=plan20.pool20,
        k=3,
        freq=freq_rank,
        delay=delay_rank,
        rank_mode=rank_mode,
        seed=seed,
    )
    s17 = sorted([n for n in plan20.pool20 if n not in excluded3])
    return Aposta17Plan(s17=s17, excluded3=sorted(excluded3), pool20=plan20.pool20)

def compute_aposta17_gate_stats(
    draws: list,
    padrao: str,
    rank_mode: str,
    window: int,
    teimosinha_n: int,
    min_hits: int,
    gap_percentis=(25.0, 75.0),
    sample_last: int = 400,
    seed: int = 1,
    metric: str = "concursos",
):
    if len(draws) < 5:
        return {"win_rate": 0.0, "wins": 0, "samples": 0, "gap_now": 0, "p_low": 0, "p_high": 0, "pass": False, "reason":"poucos concursos"}

    bases = draws[:-1]
    if sample_last and len(bases) > sample_last:
        bases = bases[-sample_last:]

    success_positions = []
    wins = 0
    samples = 0

    for base in bases:
        idx_base = base.idx  # Draw.idx já existe no script
        if idx_base <= 0:
            continue
        hist = draws[:idx_base]
        plan = build_aposta17_for_base(
            base=base,
            history_until_base=hist,
            padrao=padrao,
            rank_mode=rank_mode,
            window=window,
            seed=seed,
        )
        future = draws[idx_base + 1 : idx_base + 1 + teimosinha_n]
        if not future:
            continue

        ok = False
        for t in future:
            k = len(set(plan.s17) & set(t.nums15))
            if k >= min_hits:
                ok = True
                break

        samples += 1
        if ok:
            wins += 1
            success_positions.append(idx_base)

    gaps = [b - a for a, b in zip(success_positions, success_positions[1:])]
    p_low = _percentile(gaps, float(gap_percentis[0])) if gaps else 0.0
    p_high = _percentile(gaps, float(gap_percentis[1])) if gaps else 0.0

    gap_now = 0
    if success_positions:
        gap_now = (len(draws) - 1) - success_positions[-1]

    if not gaps:
        return {
            "win_rate": (wins / samples) if samples else 0.0,
            "wins": wins, "samples": samples,
            "gap_now": gap_now, "p_low": p_low, "p_high": p_high,
            "pass": False,
            "reason": "histórico insuficiente para calcular gaps",
            "metric": metric,
        }

    gate_pass = (gap_now >= p_low) and (gap_now <= p_high)
    reason = "" if gate_pass else f"gap atual fora da faixa histórica ({p_low}-{p_high} conc)"
    return {
        "win_rate": (wins / samples) if samples else 0.0,
        "wins": wins, "samples": samples,
        "gap_now": gap_now, "p_low": p_low, "p_high": p_high,
        "pass": gate_pass, "reason": reason, "metric": metric,
    }

def _score_candidate_nums(nums_set, freq, delay, rank_mode, window=40):
    s = 0.0
    for n in nums_set:
        # Nesta base, _score_number() aceita apenas 4 argumentos posicionais.
        # O efeito da janela já está embutido em freq/delay (rankings calculados com window).
        s += float(_score_number(n, freq, delay, rank_mode))
    return s

def _generate_top6_for_day_n(
    base_draw: Draw,
    history_until_base: list,
    padrao: str,
    rank_mode: str,
    window: int,
    seed: int,
    candidates: int,
    top6_size: int,
    overlap_target: int,
    overlap_penalty: float,
    n_nums: int,
):
    freq = _rank_frequency(history_until_base, window=window)
    delay = _rank_delay(history_until_base, window=window)

    picked = []
    picked_sets = []
    for j in range(top6_size * 3):
        use_seed = (seed + j) if (seed is not None) else None
        if n_nums == 16:
            plan = build_aposta16_for_base(
                base=base_draw,
                history_until_base=history_until_base,
                padrao=padrao,
                rank_mode=rank_mode,
                window=window,
                seed=use_seed,
            )
            nums = set(plan.s16)
        else:
            plan = build_aposta17_for_base(
                base=base_draw,
                history_until_base=history_until_base,
                padrao=padrao,
                rank_mode=rank_mode,
                window=window,
                seed=use_seed,
            )
            nums = set(plan.s17)

        score = _score_candidate_nums(nums, freq, delay, rank_mode, window=window)
        if overlap_penalty and picked_sets:
            for prev in picked_sets:
                k = len(nums & prev)
                if k >= overlap_target:
                    score -= overlap_penalty * (k - overlap_target + 1)

        # aceita se é novo ou melhor
        if nums in picked_sets:
            continue
        picked.append((score, nums))
        picked.sort(key=lambda t: t[0], reverse=True)
        picked = picked[:top6_size]
        picked_sets = [st for _, st in picked]
        if len(picked) >= top6_size:
            continue

    rows = []
    for i, (sc, st) in enumerate(picked, 1):
        rows.append({
            "rank": i,
            "base_concurso": base_draw.concurso,
            "base_data": base_draw.data,
            "nums": _fmt_set(st),
            "score": round(float(sc), 6),
            "n_nums": n_nums,
        })
    return rows

def generate_top6_today_n(
    draws: list,
    n_nums: int,
    padrao: str,
    rank_mode: str,
    window: int,
    top6_candidates: int,
    seed: int,
    top6_size: int,
    overlap_target: int,
    overlap_penalty: float,
    out_prefix: str,
    now_ts: str,
):
    base = draws[-1]
    hist = draws[:-1]
    rows = _generate_top6_for_day_n(
        base_draw=base,
        history_until_base=hist,
        padrao=padrao,
        rank_mode=rank_mode,
        window=window,
        candidates=top6_candidates,
        seed=seed,
        top6_size=top6_size,
        overlap_target=overlap_target,
        overlap_penalty=overlap_penalty,
        n_nums=n_nums,
    )
    out_csv = f"{out_prefix}_top6_s{n_nums}_{now_ts}.csv"
    if rows:
        with open(out_csv, "w", encoding="utf-8", newline="") as f:
            wr = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            wr.writeheader()
            wr.writerows(rows)
    return out_csv, rows
# =========================
# Simulação ABCD (15 dezenas) + Gate
# =========================

APOSTA15_CUSTO_DEFAULT: float = 3.00  # Ajustável via CLI (--abcd_custo15)

# Reajuste de custo da aposta (15 dezenas) a partir de 10/07/2025 (R$ 3,50).
# Regra prática: o prêmio de 11 acertos é sempre o dobro do valor da aposta.
APOSTA15_CHANGE_DATE = date(2025, 7, 10)
APOSTA15_CUSTO_ATE_20250709 = 3.00
APOSTA15_CUSTO_DESDE_20250710 = 3.50

def _infer_aposta15_custo(draw: "Draw", fallback: float = APOSTA15_CUSTO_DEFAULT) -> float:
    """Inferir custo de 1 aposta de 15 dezenas para um concurso.

    Prioridade:
      1) Se houver prêmio de 11 acertos no próprio concurso, custo = premio11 / 2.
      2) Caso contrário, usa a mudança oficial por data (10/07/2025).
      3) Fallback para o valor informado via CLI (ou default).
    """
    try:
        p11 = float(draw.premios.get(11, 0.0)) if getattr(draw, "premios", None) else 0.0
        if p11 > 0:
            # prêmio de 11 é sempre o dobro da aposta
            return round(p11 / 2.0, 2)
    except Exception:
        pass
    try:
        d = getattr(draw, "data", None) or getattr(draw, "date", None)
        if isinstance(d, date):
            return APOSTA15_CUSTO_DESDE_20250710 if d >= APOSTA15_CHANGE_DATE else APOSTA15_CUSTO_ATE_20250709
    except Exception:
        pass
    try:
        return float(fallback)
    except Exception:
        return float(APOSTA15_CUSTO_DEFAULT)

def _infer_aposta16_custo(draw: "Draw", fallback15: float = APOSTA15_CUSTO_DEFAULT) -> float:
    """Infere o custo de uma aposta de 16 números no concurso do *draw*.
    Regra: custo16 = 16 * custo15(concurso), onde custo15 é inferido via prêmio de 11 (se disponível) ou fallback por data.
    """
    return float(_infer_aposta15_custo(draw, fallback=fallback15) * 16.0)


def _infer_aposta17_custo(draw: "Draw", fallback15: float = APOSTA15_CUSTO_DEFAULT) -> float:
    """Infere o custo de uma aposta de 17 números no concurso do *draw*.
    Regra: custo17 = 136 * custo15(concurso).
    """
    return float(_infer_aposta15_custo(draw, fallback=fallback15) * 136.0)

def _build_abcd_games_from_history(history: List[Draw], janela_recente: int = 40, A_por_s16: bool = False, s16_nums: Optional[Set[int]] = None) -> Dict[str, List[int]]:
    """
    Gera os 4 jogos AB/AC/AD/BCD (15 dezenas) a partir do histórico (walk-forward).
    - Usa a mesma lógica do bloco 'ANALISE COMPLEMENTAR (OVERLAP 8/9/10)'.
    - 'history' deve conter pelo menos 1 concurso.
    - Se A_por_s16=True e s16_nums fornecido, o grupo A é escolhido para maximizar overlap com o S16 (mesma ideia do script).
    """
    if not history:
        raise ValueError("history vazio para gerar ABCD")

    last_draw = history[-1]
    # ranks / grupos baseados no concurso anterior ao alvo
    delay_rank = _rank_delay(history, window=None)
    freq_rank  = _rank_frequency(history, window=None)
    # delay_rank e freq_rank podem ser dicts ou listas; normaliza para lista ordenada por score desc
    def _topn(rank_obj, n=10):
        if isinstance(rank_obj, dict):
            # dict: num -> score (maior melhor)
            return [k for k, _ in sorted(rank_obj.items(), key=lambda kv: kv[1], reverse=True)[:n]]
        # lista/tupla de nums
        return list(rank_obj)[:n]

    B = set(_topn(delay_rank, 10))   # mais atrasados
    C = set(_topn(freq_rank, 10))    # mais frequentes

    # D = 10 mais ausentes nos últimos 'janela_recente' concursos
    recent = history[-janela_recente:] if len(history) >= janela_recente else history
    freq_recent: Dict[int, int] = {n: 0 for n in range(1, 26)}
    for d in recent:
        for n in d.bolas:
            freq_recent[n] += 1
    D = set([n for n, _ in sorted(freq_recent.items(), key=lambda kv: kv[1])[:10]])

    # A: grupos (A_global) ou escolhido por S16
    A_global, A_groups = _calc_recent_overlap_stats(history, window=janela_recente)

    if A_por_s16 and s16_nums:
        # escolhe a melhor A dentre grupos para maximizar overlap com S16
        bestA = None
        bestK = -1
        for g in A_groups:
            k = len(set(g) & set(s16_nums))
            if k > bestK:
                bestK = k
                bestA = g
        A = set(bestA) if bestA else set(A_global)
    else:
        A = set(A_global)

    jogo_A_B = sorted(A | B)
    jogo_A_C = sorted(A | C)
    jogo_A_D = sorted(A | D)
    jogo_B_C_D = sorted((B | C | D) - A)

    # sanity: todos 15
    for name, jogo in [("A_B", jogo_A_B), ("A_C", jogo_A_C), ("A_D", jogo_A_D), ("B_C_D", jogo_B_C_D)]:
        if len(jogo) != 15:
            # garante tamanho 15 ajustando (fallback simples)
            universo = list(range(1, 26))
            s = set(jogo)
            if len(s) > 15:
                jogo = sorted(list(s))[:15]
            else:
                for n in universo:
                    if n not in s:
                        s.add(n)
                        if len(s) == 15:
                            break
                jogo = sorted(list(s))
            if name == "A_B": jogo_A_B = jogo
            if name == "A_C": jogo_A_C = jogo
            if name == "A_D": jogo_A_D = jogo
            if name == "B_C_D": jogo_B_C_D = jogo

    return {
        "jogo_A_B": jogo_A_B,
        "jogo_A_C": jogo_A_C,
        "jogo_A_D": jogo_A_D,
        "jogo_B_C_D": jogo_B_C_D,
    }


def compute_abcd_gate_stats(
    draws: List[Draw],
    janela_recente: int = 40,
    teimosinha_n: int = 2,
    min_hits: int = 11,
    custo15: float = APOSTA15_CUSTO_DEFAULT,
    gate_percentis: Tuple[float, float] = (40.0, 60.0),
    metric: str = "concursos",
) -> Dict[str, Any]:
    """
    Backtest walk-forward do ABCD e calcula gate com base em gaps (entre sucessos).
    Sucesso = lucro > 0 dentro de até 'teimosinha_n' concursos jogando os 4 jogos ABCD.
    """
    if metric != "concursos":
        raise ValueError("ABCD gate: metric suportada apenas 'concursos'")

    if len(draws) < 10:
        return {"rows": [], "gaps": [], "gate": {"pass": False, "reason": "Poucos concursos"}}

    rows: List[Dict[str, Any]] = []
    success_idx: List[int] = []

    # começa em 1 porque precisa de histórico (pelo menos 1 concurso anterior)
    for i in range(1, len(draws) - (teimosinha_n - 1)):
        history = draws[:i]          # até concurso i-1
        games = _build_abcd_games_from_history(history, janela_recente=janela_recente)

        total_cost = 0.0  # custo pode variar por concurso (mudança de preço)
        total_payout = 0.0
        best_hits = 0
        win_at = None

        for r in range(teimosinha_n):
            alvo = draws[i + r]
            # custo por concurso: 4 apostas (AB/AC/AD/BCD) vezes o custo vigente no concurso
            custo15_r = _infer_aposta15_custo(alvo, fallback=custo15)
            total_cost += 4 * custo15_r
            alvo_set = set(alvo.bolas)
            # soma payout das 4 apostas
            payout_r = 0.0
            best_hits_r = 0
            for jogo in games.values():
                k = len(set(jogo) & alvo_set)
                best_hits_r = max(best_hits_r, k)
                if k >= min_hits:
                    payout_r += float(payout_for_hits(alvo, k))
            total_payout += payout_r
            best_hits = max(best_hits, best_hits_r)
            if win_at is None and payout_r > 0:
                win_at = r + 1

        profit = total_payout - total_cost
        ok = profit > 0.0

        if ok:
            success_idx.append(i)

        rows.append({
            "alvo_data": draws[i].date,
            "teimosinha_n": teimosinha_n,
            "min_hits": min_hits,
            "custo15": round(_infer_aposta15_custo(draws[i], fallback=custo15), 2),
            "custo_total": round(total_cost, 2),
            "payout_total": round(total_payout, 2),
            "profit": round(profit, 2),
            "PASS": ok,
            "win_at": win_at,
            "best_hits": best_hits,
        })

    # gaps entre sucessos
    gaps = []
    for a, b in zip(success_idx[:-1], success_idx[1:]):
        gaps.append(b - a)

    if not gaps:
        return {
            "rows": rows,
            "gaps": gaps,
            "gate": {"pass": False, "reason": "Sem sucessos suficientes para calcular percentis"},
        }

    p_low, p_high = gate_percentis
    lo = float(_percentile(gaps, p_low))
    hi = float(_percentile(gaps, p_high))

    # gap atual: desde o último sucesso até o "último dia elegível"
    last_eval_idx = len(draws) - teimosinha_n
    last_succ = max([idx for idx in success_idx if idx <= last_eval_idx], default=None)
    gap_now = (last_eval_idx - last_succ) if last_succ is not None else float("inf")

    gate_pass = (gap_now >= lo and gap_now <= hi)

    return {
        "rows": rows,
        "gaps": gaps,
        "gate": {
            "metric": metric,
            "percentis": gate_percentis,
            "faixa": (round(lo, 2), round(hi, 2)),
            "gap_atual": gap_now if gap_now != float("inf") else None,
            "pass": gate_pass,
        },
            "summary": {
        "dias_avaliados": len(rows),
        "sucessos": len(success_idx),
        "taxa": (len(success_idx) / len(rows)) if rows else 0.0,

        # Gate-level accounting (operacional)
        "dias_pass": sum(1 for r in rows if r.get("PASS")),
        "dias_skip": sum(1 for r in rows if not r.get("PASS")),
        "sucessos_pass": sum(1 for r in rows if r.get("PASS") and r.get("profit", 0.0) > 0),
        "winrate_pass": (
            (sum(1 for r in rows if r.get("PASS") and r.get("profit", 0.0) > 0) /
             max(1, sum(1 for r in rows if r.get("PASS"))))
            if rows else 0.0
        ),

        # Somas (no período)
        "custo_total_all": float(sum(r.get("custo_total", 0.0) for r in rows)) if rows else 0.0,
        "payout_total_all": float(sum(r.get("payout_total", 0.0) for r in rows)) if rows else 0.0,
        "profit_total_all": float(sum(r.get("profit", 0.0) for r in rows)) if rows else 0.0,

        "custo_total_pass": float(sum(r.get("custo_total", 0.0) for r in rows if r.get("PASS"))) if rows else 0.0,
        "payout_total_pass": float(sum(r.get("payout_total", 0.0) for r in rows if r.get("PASS"))) if rows else 0.0,
        "profit_total_pass": float(sum(r.get("profit", 0.0) for r in rows if r.get("PASS"))) if rows else 0.0,

        # Médias
        "profit_medio": float(statistics.mean([r["profit"] for r in rows])) if rows else 0.0,
        "profit_medio_pass": float(statistics.mean([r["profit"] for r in rows if r["PASS"]])) if any(r["PASS"] for r in rows) else 0.0,
    }
}



def abcd_daily_signal(
    draws: List[Draw],
    janela_recente: int,
    teimosinha_n: int,
    min_hits: int,
    custo15: float,
    gate_percentis: Tuple[float, float],
) -> Dict[str, Any]:
    """Calcula o gate (PASS/FAIL) e gera os 4 jogos ABCD para o próximo concurso.

    Obs: não roda backtest completo por dia; apenas calcula o gate com a mesma lógica
    do modo --simular_abcd_gate (baseado em gaps entre sucessos no histórico) e gera
    os jogos AB/AC/AD/BCD a partir do histórico mais recente.
    """
    stats = compute_abcd_gate_stats(
        draws=draws,
        janela_recente=janela_recente,
        teimosinha_n=teimosinha_n,
        min_hits=min_hits,
        custo15=custo15,
        gate_percentis=gate_percentis,
        metric="concursos",
    )

    jogos = _build_abcd_games_from_history(draws, janela_recente=janela_recente)
    # ordena e formata para saída
    jogos_fmt = {k: _fmt_nums(v) for k, v in jogos.items()}

    # pega contexto do último concurso conhecido
    last = draws[-1] if draws else None
    last_concurso = getattr(last, "concurso", None)
    last_data = getattr(last, "date", getattr(last, "data", None))

    gate = stats.get("gate", {}) or {}
    return {
        "last_concurso": last_concurso,
        "last_data": str(last_data) if last_data is not None else "",
        "gate": gate,
        "gate_pass": bool(gate.get("pass", False)),
        "jogos": jogos_fmt,
    }


def simulate_abcd_gate(
    draws: List[Draw],
    janela_recente: int = 40,
    teimosinha_n: int = 2,
    min_hits: int = 11,
    custo15: float = APOSTA15_CUSTO_DEFAULT,
    gate_percentis: Tuple[float, float] = (40.0, 60.0),
    out_prefix: str = "simulacao",
) -> Tuple[str, Dict[str, Any]]:
    """
    Executa compute_abcd_gate_stats e grava CSV com o backtest (PASS/profit etc).
    """
    stats = compute_abcd_gate_stats(
        draws=draws,
        janela_recente=janela_recente,
        teimosinha_n=teimosinha_n,
        min_hits=min_hits,
        custo15=custo15,
        gate_percentis=gate_percentis,
        metric="concursos",
    )

    rows = stats.get("rows", [])
    now_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_csv = f"{out_prefix}_abcd_gate_{now_ts}.csv"
    if rows:
        import csv
        with open(out_csv, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader()
            for r in rows:
                rr = dict(r)
                # data como dd/mm/yyyy
                if isinstance(rr.get("alvo_data"), date):
                    rr["alvo_data"] = rr["alvo_data"].strftime("%d/%m/%Y")
                w.writerow(rr)

    summary_src = stats.get("summary", {}) or {}

    summary = {
        "dias_avaliados": summary_src.get("dias_avaliados", len(rows)),
        "sucessos": summary_src.get("sucessos", 0),
        "taxa": summary_src.get("taxa", 0.0),
        "gate": stats.get("gate", {}),
        "csv": out_csv,

        # Operacional / financeiro
        "dias_pass": summary_src.get("dias_pass", 0),
        "dias_skip": summary_src.get("dias_skip", 0),
        "winrate_pass": summary_src.get("winrate_pass", 0.0),

        "custo_total_pass": summary_src.get("custo_total_pass", 0.0),
        "payout_total_pass": summary_src.get("payout_total_pass", 0.0),
        "profit_total_pass": summary_src.get("profit_total_pass", 0.0),

        "custo_total_all": summary_src.get("custo_total_all", 0.0),
        "payout_total_all": summary_src.get("payout_total_all", 0.0),
        "profit_total_all": summary_src.get("profit_total_all", 0.0),

        # Médias (mantém compatibilidade com prints antigos)
        "profit_medio": summary_src.get("profit_medio", 0.0),
        "profit_medio_pass": summary_src.get("profit_medio_pass", 0.0),
    }
    return out_csv, summary

def main():
    ap = argparse.ArgumentParser(description="Lotofácil: Fechamento Simples + Pool20 + Simuladores (walk-forward / ciclos)")

    # arquivo/aba
    ap.add_argument("--resultados_xlsx", default=None, help="XLSX de resultados. Se não informar, usa download/cache automático resultados_DDMMYYYY.xlsx")
    ap.add_argument("--aba", default=None, help="(opcional) Nome da aba. Se não informar, tenta LOTOFÁCIL e depois a 1ª aba.")
    ap.add_argument("--diagnostico", action="store_true", help="Mostra diagnóstico da leitura do XLSX")

    # modo
    ap.add_argument("--modo", default="fechamento", choices=["fechamento","pool20","aposta16","aposta17"])

    # fechamento
    ap.add_argument("--fix_s_mode", default="freq_window", choices=["random", "freq_window", "freq_all", "delay_window", "delay_all"],
                    help="(fechamento) Modo p/ escolher 3 fixas SORTEADAS (entre as 15 do concurso base)")
    ap.add_argument("--fix_n_mode", default="delay_window", choices=["random", "freq_window", "freq_all", "delay_window", "delay_all"],
                    help="(fechamento) Modo p/ escolher 2 fixas NÃO SORTEADAS (entre as 10 que não saíram)")

    # pool20
    ap.add_argument("--pool20_padrao", default="resultado", choices=["resultado", "moldura", "metade", "paridade"],
                    help="(pool20) Como excluir 5 dezenas para formar o pool20")
    ap.add_argument("--pool20_rank", default="mixed", choices=["freq", "delay", "mixed"],
                    help="(pool20) Ranking interno para escolher o que excluir dentro dos subconjuntos")

    # shared tuning
    ap.add_argument("--window", type=int, default=200, help="Janela de concursos para ranking (default=200). 0=usa histórico todo.")
    ap.add_argument("--seed", type=int, default=42, help="Seed determinística (default=42)")

    # ações
    ap.add_argument("--mostrar_ultimo", action="store_true", help="Mostra prévia do último concurso (alvo=último, base=penúltimo)")
    ap.add_argument("--simular", action="store_true", help="Roda simulação walk-forward em todo histórico (do 2º concurso em diante)")
    ap.add_argument("--simular_ciclos", action="store_true", help="Roda simulação de ciclos (estratégia esperar/depois jogar janela)")
    ap.add_argument("--simular_top6_gate", action="store_true",
                help="Simulação walk-forward: a cada concurso gera TOP6 (S16) usando apenas o passado e avalia se o TOP6 teria dado >=12 em até N teimosinha. Gera resumo e recomendação.")
    ap.add_argument("--top6_candidates", type=int, default=250,
                help="Quantos candidatos S16 gerar por dia (variando seed) antes de rankear e escolher o TOP6 (default=250).")
    ap.add_argument("--top6_size", type=int, default=6,
                help="Tamanho do TOP por dia (default=6).")
    ap.add_argument("--top6_teimosinha_n", type=int, default=2,
                help="Quantos concursos após o dia avaliar como teimosinha (0=apenas o próximo) (default=2).")
    ap.add_argument("--top6_eval_min_hits", type=int, default=12,
                help="Mínimo de acertos na aposta16 para considerar sucesso (default=12).")
    ap.add_argument("--top6_recent_window", type=int, default=50,
                help="Janela (concursos) para métricas e ranking diário (freq/delay). Se 0, usa histórico todo (default=50).")
    ap.add_argument("--top6_overlap_target", type=int, default=9,
                help="Alvo de overlap com o concurso base (default=9). Penaliza desvio.")
    ap.add_argument("--top6_overlap_penalty", type=float, default=35.0,
                help="Peso da penalização por |k_base - alvo| no score diário (default=35.0).")
    ap.add_argument("--top6_gate_percentis", default="40,60",
                help="Faixa de percentis (low,high) para o gate de gap (default=40,60). Ex: 20,80")

    # ABCD gate (4 jogos de 15 dezenas)
    ap.add_argument("--simular_abcd_gate", action="store_true",
                    help="Simula walk-forward do AB/AC/AD/BCD (4 jogos de 15) e calcula gate (PASS/FAIL) por gaps em concursos.")
    ap.add_argument("--abcd_janela_recente", type=int, default=40,
                    help="Janela recente (concursos) usada para calcular grupos/ausências no ABCD (default: 40).")
    ap.add_argument("--abcd_teimosinha_n", type=int, default=2,
                    help="Teimosinha para ABCD: repete os 4 jogos por N concursos no backtest (default: 2).")
    ap.add_argument("--abcd_min_hits", type=int, default=11,
                    help="Hits mínimos para considerar payout no ABCD (default: 11).")
    ap.add_argument("--abcd_custo15", type=float, default=APOSTA15_CUSTO_DEFAULT,
                    help="Custo de 1 aposta de 15 dezenas (R$). O backtest usa 4 apostas por concurso. (default: 3.00).")
    ap.add_argument("--abcd_gate_percentis", type=str, default="40,60",
                    help="Percentis (p_low,p_high) para faixa do gate do ABCD (default: 40,60).")

    ap.add_argument("--abcd_daily_signal", action="store_true",
                    help="Gera sinal diário (gate PASS/FAIL) e imprime os 4 jogos ABCD sugeridos para o próximo concurso (sem rodar o backtest completo).")
    ap.add_argument("--abcd_daily_json", type=str, default="",
                    help="Se informado, grava um JSON com o sinal diário e os 4 jogos ABCD neste caminho.")
    ap.add_argument("--top6_gate_min_trials", type=int, default=120,
                help="Mínimo de gaps (entre sucessos) para calcular gate (default=120).")

    # cartões e custos
    
    ap.add_argument("--repeats_min", type=int, default=2,
                    help="Gera relatório de jogos repetidos (mesmos 15 números no mesmo cartão) quando um jogo é gerado pelo menos N vezes. Use 1 para listar todos.")
    
    ap.add_argument("--gerarapostas", action="store_true",
                    help="Gera um relatório TOP (engenharia reversa) a partir do CSV de repetidos: filtra jogos com best_hits_today>=12 e rankeia por frequência e menor intervalo.")
    ap.add_argument("--gerarapostas_top", type=int, default=6,
                    help="Quantos jogos sugeridos listar (default=6). Regra: 1 jogo por dia (origem).")
    ap.add_argument("--gerarapostas_min_hits", type=int, default=12,
                    help="Filtro mínimo de best_hits_today para considerar no TOP (default=12).")
    ap.add_argument("--gerarapostas_repetidos_csv", default=None,
                    help="(opcional) Caminho do CSV de repetidos. Se não informar, tenta achar o mais recente 'simulacao_repetidos_*.csv' no diretório.")
    ap.add_argument("--gerar_top6", action="store_true",
                    help="(gerarapostas) Em vez de usar o CSV de repetidos (TOP #1-6), gera TOP6 (6 jogos S16) com base no ÚLTIMO concurso do XLSX.")
    ap.add_argument("--gate_top6", action="store_true",
                    help="(gerarapostas + --gerar_top6) Aplica gate do TOP6 (walk-forward) usando métrica em concursos (recomendado).")
    ap.add_argument("--top6_gate_window", type=int, default=400,
                    help="(gate_top6) Quantos concursos-base usar para estimar percentis do gate TOP6. Default=400.")
    ap.add_argument("--janela_recente", type=int, default=40,
                    help="(gerarapostas) Janela recente (30-50 recomendado) para analise de overlap e scores. Default=40.")

    ap.add_argument("--A_por_s16", action="store_true",
                    help="(gerarapostas) Se ativo, escolhe o Grupo A (10 dezenas do último concurso) condicionado ao S16 (prioriza S16 ∩ último resultado).")
    # gate (aposta16): decide se vale a pena gerar/jogar hoje baseado em histórico de >=12 dentro de (1+teimosinha) concursos
    ap.add_argument("--gate_aposta16", action="store_true",
                    help="(aposta16/gerarapostas) Ativa um 'gate' que só sugere jogos se o gap atual (dias desde o último >=min_hits) estiver dentro da faixa histórica.")
    ap.add_argument("--gate_min_hits", type=int, default=12,
                    help="(gate) Hits mínimos para considerar 'sucesso' no backtest (default=12).")
    ap.add_argument("--gate_teimosinha_n", type=int, default=2,
                    help="(gate) Quantos concursos adicionais testar após o próximo (teimosinha). Ex: 2 => testa 3 concursos (next +2). Default=2.")
    ap.add_argument("--gate_lookback", type=int, default=400,
                    help="(gate) Quantos concursos-base usar no backtest (default=400).")
    ap.add_argument("--gate_gap_percentis", type=str, default="30,70",
                    help="(gate) Percentis p/ faixa de gap em dias. Ex: 25,75. Default=30,70.")
    ap.add_argument("--gate_min_win_rate", type=float, default=0.0,
                    help="(gate) Win-rate mínimo no backtest para permitir sugestão (0..1). Default=0.0 (desliga).")


    ap.add_argument("--usar_cartoes", default=None, help="Quais cartões considerar como 'jogados'. Ex: AS,BS ou P1,P2")
    ap.add_argument("--custo_por_cartao", type=float, default=3.0, help="Custo por cartão (R$) para calcular ROI. (Ignorado no modo aposta16; usa R$ 56,00 hardcoded.) Default=3.00")

    # saída
    ap.add_argument("--saida_prefixo", default="simulacao", help="Prefixo do CSV. Arquivo final inclui modo/padrão + timestamp")
    # Flags de geração (novas)
    ap.add_argument("--gerar_top6_s16", action="store_true")
    ap.add_argument("--gerar_top6_s17", action="store_true")
    ap.add_argument("--gerar_s16", action="store_true")
    ap.add_argument("--gerar_s17", action="store_true")
    ap.add_argument("--gerar_abcd", action="store_true")

    # ciclos
    ap.add_argument("--ciclos_waits", type=str, default=None,
                    help="Lista de waits (inteiros) separados por vírgula. Ex: 8,9,10")
    ap.add_argument("--ciclos_waits_loss", type=str, default=None,
                    help="Lista de waits (perdas consecutivas sem >=11) para disparar a aposta. Ex: 8,9,10")
    ap.add_argument("--ciclos_janelas", type=str, default=None,
                    help="Lista de janelas (inteiros) separados por vírgula. Ex: 1,2")
    ap.add_argument("--ciclos_teimosinha", action="store_true", help="(ciclos) Ativa teimosinha simples (se janela=1, joga 2 concursos)")
    ap.add_argument("--ciclos_teimosinha_n", type=int, choices=[0,1,2], default=0,
                    help="(ciclos) Teimosinha: repete o mesmo jogo por +N concursos após o gatilho (0..2). Se usar --ciclos_teimosinha, equivale a 1.")
    ap.add_argument("--ciclos_periodico", action="store_true", help="(ciclos) Após fim da janela, volta para WAIT mesmo sem ganho")
    ap.add_argument("--ciclos_detalhar", type=str, default=None, help="Ex: 8:1 para gerar detalhe só para wait=8, janela=1")

    args = ap.parse_args()

    resultados_path = ensure_results_file(args.resultados_xlsx)

    draws, sheet_used = read_draws_xlsx(resultados_path, sheet_name=args.aba, diagnostico=args.diagnostico)
    print(f"Resultados XLSX: {resultados_path} | Aba: {sheet_used}")

    # atualiza ResultadoNorm para facilitar validação/search
    ensure_resultado_norm_column(resultados_path, args.aba, draws, sheet_used)
    # ===== Atalhos de geração (flags específicas) =====
    now_ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    if args.gerar_top6_s16 or args.gerar_top6_s17:
        n_nums = 16 if args.gerar_top6_s16 else 17
        print(f"\n=== GERAR TOP6 S{n_nums} (hoje) ===")
        out_csv, rows = generate_top6_today_n(
            draws=draws,
            n_nums=n_nums,
            padrao=args.pool20_padrao,
            rank_mode=args.pool20_rank,
            window=args.janela_recente,
            top6_candidates=args.top6_candidates,
            seed=args.seed,
            top6_size=args.top6_size,
            overlap_target=args.top6_overlap_target,
            overlap_penalty=args.top6_overlap_penalty,
            out_prefix=args.saida_prefixo,
            now_ts=now_ts,
        )
        for r in rows:
            print(f"#{r['rank']} | S{n_nums} | {r['nums']} | score={r['score']}")
        print(f"CSV: {out_csv}")
        return

    if args.gerar_s16:
        print("\n=== GERAR S16 (1 jogo) ===")
        base = draws[-1]
        hist = draws[:-1]
        plan16 = build_aposta16_for_base(
            base=base,
            history_until_base=hist,
            padrao=args.pool20_padrao,
            rank_mode=args.pool20_rank,
            window=args.janela_recente,
            seed=args.seed,
        )
        print(f"S16: {_fmt_set(plan16.s16)} | excluídas(5): {_fmt_list(plan16.excluded5)} | excluídas(4): {_fmt_list(plan16.excluded4)}")

        if args.gate_aposta16:
            gate = compute_aposta16_gate_stats(
                draws=draws,
            n_nums=(17 if args.modo=='aposta17' else 16),
            padrao=args.pool20_padrao,
                rank_mode=args.pool20_rank,
                window=args.janela_recente,
                teimosinha_n=args.gate_teimosinha_n,
                min_hits=args.gate_min_hits,
                gap_percentis=tuple(args.gate_gap_percentis),
                sample_last=400,
                seed=args.seed,
                metric="concursos",
            )
            print(f"[GATE S16] win_rate={gate['win_rate']:.3f} ({gate['wins']}/{gate['samples']}) | gap_atual={gate['gap_now']}conc | faixa={gate['p_low']}-{gate['p_high']}conc | PASS={gate['pass']}")
            if not gate["pass"]:
                print("-> Gate não passou: não recomendo jogar hoje.")
        return

    if args.gerar_s17:
        print("\n=== GERAR S17 (1 jogo) ===")
        base = draws[-1]
        hist = draws[:-1]
        plan17 = build_aposta17_for_base(
            base=base,
            history_until_base=hist,
            padrao=args.pool20_padrao,
            rank_mode=args.pool20_rank,
            window=args.janela_recente,
            seed=args.seed,
        )
        print(f"S17: {_fmt_set(plan17.s17)} | excluídas(3): {_fmt_list(plan17.excluded3)}")

        if args.gate_aposta17:
            gate = compute_aposta17_gate_stats(
                draws=draws,
                padrao=args.pool20_padrao,
                rank_mode=args.pool20_rank,
                window=args.janela_recente,
                teimosinha_n=args.gate_teimosinha_n,
                min_hits=args.gate_min_hits,
                gap_percentis=tuple(args.gate_gap_percentis),
                sample_last=400,
                seed=args.seed,
                metric="concursos",
            )
            print(f"[GATE S17] win_rate={gate['win_rate']:.3f} ({gate['wins']}/{gate['samples']}) | gap_atual={gate['gap_now']}conc | faixa={gate['p_low']}-{gate['p_high']}conc | PASS={gate['pass']}")
            if not gate["pass"]:
                print("-> Gate não passou: não recomendo jogar hoje.")
        return

    if args.gerar_abcd:
        print("\n=== GERAR AB/AC/AD/BCD ===")
        overlap_csv = find_latest_file(f"{args.saida_prefixo}_sugeridas_overlap_*.csv")
        if overlap_csv:
            rows = load_csv_rows(overlap_csv)
            if rows:
                r0 = rows[0]
                for k in ["jogo_A_B","jogo_A_C","jogo_A_D","jogo_B_C_D"]:
                    if k in r0 and r0[k]:
                        print(f"{k}: {r0[k]}")
                print(f"OVERLAP CSV: {overlap_csv}")
                return
        print("[ERRO] Não encontrei CSV de overlap. Rode --gerarapostas primeiro para gerar os relatórios.")
        return

    modo = args.modo.strip().lower()

    # default --usar_cartoes depende do modo
    if args.usar_cartoes is None:
        args.usar_cartoes = "AR,AS,BR,BS" if modo == "fechamento" else ("P1,P2,P3,P4" if modo == "pool20" else "S16")

    use_cards = [c.strip().upper() for c in (args.usar_cartoes or "").split(",") if c.strip()]
    if not use_cards:
        raise SystemExit("ERRO: --usar_cartoes vazio.")


    # STANDALONE_GERARAPOSTAS: se o usuário pediu só para gerar apostas (sem simular),
    # usa o CSV de repetidos mais recente (ou o informado) e gera o relatório TOP.
    if args.gerarapostas and (not args.simular) and (not args.simular_ciclos):
        repetidos_path = args.gerarapostas_repetidos_csv or _find_latest_repeats_csv(prefix=f"{args.saida_prefixo}_repetidos_")
        if not repetidos_path:
            repetidos_path = args.gerarapostas_repetidos_csv or _find_latest_repeats_csv(prefix="simulacao_repetidos_")
        if not repetidos_path or not os.path.exists(repetidos_path):
            raise SystemExit("ERRO: não encontrei CSV de repetidos. Rode antes com --simular e repeats_min>=2, ou informe --gerarapostas_repetidos_csv.")

        out_csv, top_rows = generate_apostas_from_repetidos(
            repetidos_csv=repetidos_path,
            top_n=int(args.gerarapostas_top),
            min_best_hits_today=int(args.gerarapostas_min_hits),
            min_times_generated=max(2, int(args.repeats_min or 2)),
            out_prefix=args.saida_prefixo,
        )

        print("\n=== GERAR APOSTAS (TOP) ===")
        print(f"Fonte repetidos: {repetidos_path}")
        print(f"Relatório TOP:   {out_csv}")
        if top_rows:
            print("\nTop sugeridos (1 por dia):")
            for r in top_rows:
                print(
                    f"#{r.get('rank')} | {r.get('origin_target_data')} | {r.get('card')} | {r.get('nums')} | "
                    f"freq={r.get('times_generated')} | best={r.get('best_hits_today')} | "
                    f"minGapConc={r.get('min_gap_concurso')} | avgGapConc={r.get('avg_gap_concurso')} | "
                    f"payoutSum={r.get('sum_payout_today')}"
                )

        # Analise complementar (overlap + estrategia A/B/C/D)
        overlap_csv, overlap_rows = generate_overlap_analysis_for_top(
                    draws=draws,
                    top_rows=top_rows,
            janela_recente=int(args.janela_recente or 0),
            A_por_s16=bool(getattr(args, 'A_por_s16', False)),
            out_prefix=args.saida_prefixo,
        )
        # Gate opcional: só recomendar apostas hoje se o gap atual (dias desde o último >=min_hits)
        # estiver dentro de uma faixa histórica calculada via backtest do próprio pipeline Aposta16.
        gate_info = None
        if bool(getattr(args, "gate_aposta16", False)) and modo == "aposta16":
            try:
                p_lo, p_hi = (30.0, 70.0)
                raw = str(getattr(args, "gate_gap_percentis", "30,70") or "30,70")
                parts = [x.strip() for x in raw.split(",") if x.strip()]
                if len(parts) >= 2:
                    p_lo, p_hi = float(parts[0]), float(parts[1])
            except Exception:
                p_lo, p_hi = (30.0, 70.0)

            gate_info = compute_aposta16_gate_stats(
                draws=draws,
                pool20_padrao=args.pool20_padrao,
                pool20_rank=args.pool20_rank,
                window=int(args.window or 0),
                seed=args.seed,
                min_hits=int(getattr(args, "gate_min_hits", 12) or 12),
                teimosinha_n=int(getattr(args, "gate_teimosinha_n", 2) or 0),
                lookback_bases=int(getattr(args, "gate_lookback", 400) or 0),
                gap_percentis=(p_lo, p_hi),
                metric="concursos",
            )

            # win-rate mínimo (opcional)
            min_wr = float(getattr(args, "gate_min_win_rate", 0.0) or 0.0)
            if gate_info and min_wr > 0.0 and float(gate_info.get("win_rate", 0.0)) < min_wr:
                gate_info["gate_pass"] = False
                gate_info["gate_reason"] = f"win-rate abaixo do mínimo ({gate_info.get('win_rate',0.0):.3f} < {min_wr:.3f})"

            # injeta no CSV e sobrescreve o arquivo OVERLAP já gerado
            if overlap_rows and overlap_csv and gate_info:
                for rr in overlap_rows:
                    rr["gate_min_hits"] = int(getattr(args, "gate_min_hits", 12) or 12)
                    rr["gate_teimosinha_n"] = int(getattr(args, "gate_teimosinha_n", 2) or 0)
                    rr["gate_lookback"] = int(getattr(args, "gate_lookback", 400) or 0)
                    rr["gate_gap_p_low"] = gate_info.get("p_low")
                    rr["gate_gap_p_high"] = gate_info.get("p_high")
                    rr["gate_current_gap_days"] = gate_info.get("current_gap_days")
                    rr["gate_win_rate"] = round(float(gate_info.get("win_rate", 0.0)), 4)
                    rr["gate_samples"] = gate_info.get("samples")
                    rr["gate_successes"] = gate_info.get("successes")
                    rr["gate_pass"] = gate_info.get("gate_pass")
                    rr["gate_reason"] = gate_info.get("gate_reason")
                try:
                    _write_csv_dicts(overlap_csv, overlap_rows)
                except Exception:
                    pass

        if overlap_csv:
            print("\n=== ANALISE COMPLEMENTAR (OVERLAP 8/9/10) ===")
            print(f"Janela recente: {int(args.janela_recente or 0)} concursos")
            print(f"Relatório OVERLAP: {overlap_csv}")
            if gate_info:
                metric = str(gate_info.get("metric", "dias") or "dias").lower()
                unit = "conc" if metric == "concursos" else "d"
                cg = gate_info.get("gap_now")
                if cg is None:
                    cg = gate_info.get("current_gap_concursos") if metric == "concursos" else gate_info.get("current_gap_days")
                try:
                    cg = int(cg)
                except Exception:
                    cg = 0
                print("\n[GATE aposta16] sucesso>= {mh} em até {tn} teimosinha | win_rate={wr:.3f} ({succ}/{samp}) | "
                      "gap_atual={cg}{unit} | faixa={pl:.1f}-{ph:.1f}{unit} | PASS={gp} | {rsn}".format(
                          mh=int(getattr(args, "gate_min_hits", 12) or 12),
                          tn=int(getattr(args, "gate_teimosinha_n", 2) or 0),
                          wr=float(gate_info.get("win_rate", 0.0)),
                          succ=int(gate_info.get("successes", 0) or 0),
                          samp=int(gate_info.get("samples", 0) or 0),
                          cg=cg,
                          unit=unit,
                          pl=float(gate_info.get("p_low", 0.0)),
                          ph=float(gate_info.get("p_high", 0.0)),
                          gp=bool(gate_info.get("gate_pass", False)),
                          rsn=str(gate_info.get("gate_reason", "")),
                      ))
                if not bool(gate_info.get("gate_pass", False)):
                    print("-> Gate não passou: não recomendo jogar hoje (consulte o CSV para detalhes).")
                    return
            if overlap_rows:
                print("\nRanking por overlap/score (melhores primeiro):")
                for rr in overlap_rows[:min(6, len(overlap_rows))]:
                    print(
                        f"#{rr.get('rank_overlap')} | k_last={rr.get('k_last_overlap')} ({rr.get('k_last_bucket')}) | "
                        f"score={rr.get('score_final')} | {rr.get('card')} | {rr.get('nums')}"
                    )

        return

    if args.mostrar_ultimo:
        print_last_preview(
            modo=modo,
            fix_s_mode=args.fix_s_mode,
            fix_n_mode=args.fix_n_mode,
            pool20_padrao=args.pool20_padrao,
            pool20_rank=args.pool20_rank,
            window=args.window,
            seed=args.seed,
        )

    if args.simular:
        print("[SIM] Executando simulação walk-forward...")
        stats, payout_total, custo_total, full_csv, repeats_csv = simulate_walk_forward(
            modo=modo,
            use_cards=use_cards,
            custo_por_cartao=args.custo_por_cartao,
            fix_s_mode=args.fix_s_mode,
            fix_n_mode=args.fix_n_mode,
            pool20_padrao=args.pool20_padrao,
            pool20_rank_mode=args.pool20_rank,
            window=args.window,
            seed=args.seed,
            out_prefix=args.saida_prefixo,
            repeats_min=args.repeats_min,
        )

        filtered_stats = {c: stats[c] for c in use_cards}
        print_sim_summary(filtered_stats, payout_total, custo_total, use_cards)

        print("\n=== ARQUIVO GERADO ===")
        print(f"Relatório completo: {full_csv}")
        if repeats_csv:
            print(f"Relatório de repetidos (repeats_min={args.repeats_min}): {repeats_csv}")


        # GERARAPOSTAS_POSTSIM: opcionalmente gerar relatório TOP baseado no repetidos recém-criado
        if args.gerarapostas:
            if not repeats_csv or (not os.path.exists(repeats_csv)):
                print("[WARN] --gerarapostas foi pedido, mas não há CSV de repetidos (repeats_min precisa ser >=2).")
            else:
                # --- (gerarapostas) opção: gerar TOP6 (6 jogos S16) baseado no último concurso ---
                if getattr(args, "gerar_top6", False):
                    # Gate TOP6 (opcional, recomendado com metric=concursos)
                    if getattr(args, "gate_top6", False):
                        try:
                            p_lo, p_hi = [float(x) for x in str(args.top6_gate_percentis).split(",")]
                        except Exception:
                            p_lo, p_hi = 40.0, 60.0

                        gate = compute_top6_gate_stats(
                            draws=draws,
                            padrao=args.pool20_padrao,
                            rank_mode=args.pool20_rank,
                            top6_candidates=int(args.top6_candidates),
                            top6_size=int(args.top6_size),
                            teimosinha_n=int(args.top6_teimosinha_n),
                            min_hits=int(args.top6_eval_min_hits),
                            gate_percentis=(p_lo, p_hi),
                            metric="concursos",
                            window_bases=int(getattr(args, "top6_gate_window", 400)),
                            overlap_penalty=float(getattr(args, "top6_overlap_penalty", 0.0)),
                        )

                        print(
                            f"\n[GATE TOP6] metric=concursos | percentis={p_lo:.1f},{p_hi:.1f} | "
                            f"faixa={gate['p_low']}-{gate['p_high']} conc | gap_atual={gate['gap_now']} conc | "
                            f"PASS={gate['pass']} | win_rate={gate['win_rate']:.3f} ({gate['success']}/{gate['total']})"
                        )
                        if not gate["pass"]:
                            msg = gate.get("reason", "") or "gate não passou"
                            print(f"-> Gate TOP6 não passou: não recomendo gerar/jogar hoje ({msg}).")
                            return

                    top6 = generate_top6_today(
                        draws=draws,
                        padrao=args.pool20_padrao,
                        rank_mode=args.pool20_rank,
                        top6_candidates=int(args.top6_candidates),
                        top6_size=int(args.top6_size),
                        overlap_penalty=float(getattr(args, "top6_overlap_penalty", 0.0)),
                    )

                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    out_csv = f"top6_sugeridas_{ts}.csv"
                    with open(out_csv, "w", encoding="utf-8", newline="") as f:
                        w = csv.writer(f, delimiter=";")
                        w.writerow(["rank", "card", "nums"])
                        for i, s16 in enumerate(top6, 1):
                            w.writerow([i, "S16", _fmt_nums(s16)])

                    print("\n=== GERAR APOSTAS (TOP6 S16) ===")
                    print(f"Relatório TOP6: {out_csv}")
                    for i, s16 in enumerate(top6, 1):
                        print(f"#{i} | S16 | {_fmt_nums(s16)}")

                    return

                out_csv, top_rows = generate_apostas_from_repetidos(
                    repetidos_csv=repeats_csv,
                    top_n=int(args.gerarapostas_top),
                    min_best_hits_today=int(args.gerarapostas_min_hits),
                    min_times_generated=max(2, int(args.repeats_min or 2)),
                    out_prefix=args.saida_prefixo,
                )
                print("\n=== GERAR APOSTAS (TOP) ===")
                print(f"Relatório TOP: {out_csv}")
                if top_rows:
                    print("\nTop sugeridos (1 por dia):")
                    for r in top_rows:
                        print(
                            f"#{r.get('rank')} | {r.get('origin_target_data')} | {r.get('card')} | {r.get('nums')} | "
                            f"freq={r.get('times_generated')} | best={r.get('best_hits_today')} | "
                            f"minGapConc={r.get('min_gap_concurso')} | avgGapConc={r.get('avg_gap_concurso')} | "
                            f"payoutSum={r.get('sum_payout_today')}"
                        )

                # Analise complementar (overlap + estrategia A/B/C/D)
                overlap_csv, overlap_rows = generate_overlap_analysis_for_top(
                    draws=draws,
                    top_rows=top_rows,
                    janela_recente=int(args.janela_recente or 0),
                    A_por_s16=bool(args.A_por_s16),
                    out_prefix=args.saida_prefixo,
                )
                if overlap_csv:
                    print("\n=== ANALISE COMPLEMENTAR (OVERLAP 8/9/10) ===")
                    print(f"Janela recente: {int(args.janela_recente or 0)} concursos")
                    print(f"Relatório OVERLAP: {overlap_csv}")
                    if overlap_rows:
                        print("\nRanking por overlap/score (melhores primeiro):")
                        for rr in overlap_rows[:min(6, len(overlap_rows))]:
                            print(
                                f"#{rr.get('rank_overlap')} | k_last={rr.get('k_last_overlap')} ({rr.get('k_last_bucket')}) | "
                                f"score={rr.get('score_final')} | {rr.get('card')} | {rr.get('nums')}"
                            )

    if args.simular_top6_gate:
        # parse percentis
        try:
            p1, p2 = [float(x.strip()) for x in str(args.top6_gate_percentis).split(",")[:2]]
        except Exception:
            p1, p2 = 40.0, 60.0
    
        print("[SIM] Executando simulação TOP6 diário + gate (walk-forward)...")
        csv_path, summary = simulate_top6_gate(
            draws=draws,
            n_nums=17 if args.modo == "aposta17" else 16,
            padrao=args.pool20_padrao,
            rank_mode=args.pool20_rank,
            window=int(args.top6_recent_window),
            seed=int(args.seed),
            top6_candidates=int(args.top6_candidates),
            top6_size=int(args.top6_size),
            teimosinha_n=int(args.top6_teimosinha_n),

            min_hits=int(args.top6_eval_min_hits),
            overlap_target=int(args.top6_overlap_target),
            overlap_penalty=float(args.top6_overlap_penalty),
            gate_percentis=(p1, p2),
            gate_min_trials=int(args.top6_gate_min_trials),
            out_prefix="simulacao",
        )
    
        print("\n=== SIMULAÇÃO TOP6 (diária) ===")
        print(f"Dias avaliados: {summary.get('days')} | sucessos(lucro>0): {summary.get('success_days')} | taxa: {summary.get('success_rate',0.0):.3f}")
        gate = summary.get("gate", {})
        if gate:
            faixa = gate.get("faixa", (None, None))
            gap_atual = gate.get("gap_atual", None)
            pcts = gate.get("percentis", (None, None))
            print(f"[GATE TOP6] metric=concursos | percentis={pcts[0]},{pcts[1]} | faixa={faixa[0]}-{faixa[1]} conc | gap_atual={gap_atual} conc | PASS={gate.get('pass', False)}")
        print(f"Dias PASS: {summary.get('dias_pass', 0)} | Dias SKIP: {summary.get('dias_skip', 0)} | Winrate PASS (lucro>0): {summary.get('winrate_pass', 0.0):.3f}")
        print(f"Totais (somente PASS): custo={summary.get('custo_total_pass', 0.0):.2f} | payout={summary.get('payout_total_pass', 0.0):.2f} | profit={summary.get('profit_total_pass', 0.0):.2f}")
        print(f"Totais (todos dias):   custo={summary.get('custo_total_all', 0.0):.2f} | payout={summary.get('payout_total_all', 0.0):.2f} | profit={summary.get('profit_total_all', 0.0):.2f}")
        print(f"Profit médio (todos dias): {summary.get('profit_medio', 0.0):.2f} | Profit médio (somente PASS): {summary.get('profit_medio_pass', 0.0):.2f}")
        print(f"Relatório: {csv_path}")
        return

    
    if args.abcd_daily_signal:
        # Sinal diário para o próximo concurso (gate + 4 jogos ABCD)
        p = _parse_percentiles(args.abcd_gate_percentis, default=(40.0, 60.0))
        sig = abcd_daily_signal(
            draws=draws,
            janela_recente=int(args.abcd_janela_recente),
            teimosinha_n=int(args.abcd_teimosinha_n),
            min_hits=int(args.abcd_min_hits),
            custo15=float(args.abcd_custo15),
            gate_percentis=p,
        )

        gate = sig.get("gate", {}) or {}
        print("\n=== SINAL DIÁRIO ABCD ===")
        print(f"Último concurso: {sig.get('last_concurso')} | Data: {sig.get('last_data')}")
        print(f"[GATE ABCD] metric={gate.get('metric','concursos')} | percentis={gate.get('percentis')} | "
              f"faixa={gate.get('faixa')} conc | gap_atual={gate.get('gap_atual')} conc | PASS={sig.get('gate_pass')}")
        print("Jogos sugeridos (15 dezenas):")
        for k in ["AB", "AC", "AD", "BCD"]:
            if k in sig["jogos"]:
                print(f"  {k}: {sig['jogos'][k]}")
        if args.abcd_daily_json:
            try:
                Path(args.abcd_daily_json).write_text(json.dumps(sig, ensure_ascii=False, indent=2), encoding="utf-8")
                print(f"JSON: {args.abcd_daily_json}")
            except Exception as e:
                print(f"[WARN] Falha ao gravar JSON em {args.abcd_daily_json}: {e}")
        return

    if args.simular_abcd_gate:
        print("[SIM] Executando simulação ABCD + gate (walk-forward)...")
        p = _parse_percentiles(args.abcd_gate_percentis, default=(40.0, 60.0))
        csv_path, summary = simulate_abcd_gate(
            draws=draws,
            janela_recente=int(args.abcd_janela_recente),
            teimosinha_n=int(args.abcd_teimosinha_n),
            min_hits=int(args.abcd_min_hits),
            custo15=float(args.abcd_custo15),
            gate_percentis=p,
            out_prefix="simulacao",
        )
        gate = summary.get("gate", {})
        print("\n=== SIMULAÇÃO ABCD (diária) ===")
        print(f"Dias avaliados: {summary['dias_avaliados']} | sucessos(lucro>0): {summary['sucessos']} | taxa: {summary['taxa']:.3f}")
        if gate:
            faixa = gate.get("faixa", (None, None))
            gap_atual = gate.get("gap_atual", None)
            pcts = gate.get("percentis", p)
            print(f"[GATE ABCD] metric=concursos | percentis={pcts[0]},{pcts[1]} | faixa={faixa[0]}-{faixa[1]} conc | gap_atual={gap_atual} conc | PASS={gate.get('pass', False)}")
        
        print(f"Dias PASS: {summary.get('dias_pass', 0)} | Dias SKIP: {summary.get('dias_skip', 0)} | Winrate PASS (lucro>0): {summary.get('winrate_pass', 0.0):.3f}")
        print(f"Totais (somente PASS): custo={summary.get('custo_total_pass', 0.0):.2f} | payout={summary.get('payout_total_pass', 0.0):.2f} | profit={summary.get('profit_total_pass', 0.0):.2f}")
        print(f"Totais (todos dias):   custo={summary.get('custo_total_all', 0.0):.2f} | payout={summary.get('payout_total_all', 0.0):.2f} | profit={summary.get('profit_total_all', 0.0):.2f}")
        print(f"Profit médio (todos dias): {summary.get('profit_medio', 0.0):.2f} | Profit médio (somente PASS): {summary.get('profit_medio_pass', 0.0):.2f}")
        print(f"Relatório: {csv_path}")
        return
    if args.simular_ciclos:
        # teimosinha_n: preferir parâmetro novo (0..2); flag antiga vira 1
        teimosinha_n = int(getattr(args, "ciclos_teimosinha_n", 0) or 0)
        if getattr(args, "ciclos_teimosinha", False) and teimosinha_n == 0:
            teimosinha_n = 1

        waits_win = _parse_int_csv(args.ciclos_waits, default=[])
        wins = _parse_int_csv(args.ciclos_janelas, default=[])
        waits_loss = _parse_int_csv(args.ciclos_waits_loss, default=[0])

        if not waits_win or not wins:
            raise SystemExit("ERRO: --ciclos_waits e --ciclos_janelas precisam ter pelo menos 1 valor cada.")

        # --ciclos_detalhar aceita:
        #   "w:pw"         (compat)  -> wait_after_win=w, wait_after_loss=0, play_window=pw
        #   "w:wl:pw"      (novo)    -> wait_after_win=w, wait_after_loss=wl, play_window=pw
        detail_filter: Set[Tuple[int, int, int]] = set()
        if args.ciclos_detalhar:
            for token in str(args.ciclos_detalhar).split(","):
                t = token.strip()
                if not t:
                    continue
                parts = [p for p in t.split(":") if p.strip()]
                if len(parts) == 2:
                    w = int(parts[0]); pw = int(parts[1]); wl = 0
                elif len(parts) == 3:
                    w = int(parts[0]); wl = int(parts[1]); pw = int(parts[2])
                else:
                    raise SystemExit("ERRO: --ciclos_detalhar inválido. Use 'w:pw' ou 'w:wl:pw'. Ex: 7:8:2")
                detail_filter.add((w, wl, pw))

        summaries: List[CycleSummary] = []
        details_all: List[Dict[str, object]] = []

        for w in waits_win:
            for wl in waits_loss:
                for pw in wins:
                    summary, detalhe = simulate_cycles_strategy(
                        modo=args.modo,
                        pool20_padrao=args.pool20_padrao,
                        pool20_rank=args.pool20_rank,
                        wait_after_win=w,
                        wait_after_loss=wl,
                        play_window=pw,
                        teimosinha_n=teimosinha_n,
                        periodic=args.ciclos_periodico,
                        use_cards=use_cards,
                        custo_por_cartao=args.custo_por_cartao,
                        fix_s_mode=args.fix_s_mode,
                        fix_n_mode=args.fix_n_mode,
                        window=args.window,
                        seed=args.seed,
                    )
                    summaries.append(summary)
                    if (not detail_filter) or ((w, wl, pw) in detail_filter):
                        details_all.extend(detalhe)

        stamp = now_stamp()
        resumo_csv = f"ciclos_resumo_{stamp}.csv"
        detalhe_csv = f"ciclos_detalhe_{stamp}.csv"
        _write_csv_dicts(resumo_csv, [s.__dict__ for s in summaries])
        _write_csv_dicts(detalhe_csv, details_all)

        print("\n=== ARQUIVOS GERADOS (ciclos) ===")
        print(f"Resumo:  {resumo_csv}")
        print(f"Detalhe: {detalhe_csv}")




if __name__ == "__main__":
    main()
