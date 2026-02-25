# abcd_runner.py
from __future__ import annotations

import argparse
import json
import re
import subprocess
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple
from datetime import datetime, date
from zoneinfo import ZoneInfo
import shutil
from openpyxl import load_workbook

# This runner delegates ALL ABCD logic (including gate calculation) to the main script,
# and only:
# - writes a daily snapshot into docs/results/YYYY/MM/YYYY-MM-DD.json
# - maintains campaign state in docs/state/abcd_campaigns.json
# - creates email_body.txt when there is something to email (new campaign, active reminders, win/expire)
#
# It is intentionally "dumb" about the gate to avoid drift vs your validated v18/v19 logic.

SCRIPT = "fechamento_simples_v11_54_abcd_gate_plus_FIXED19.py"

DOCS_DIR = Path("docs")
RESULTS_DIR = DOCS_DIR / "results"
STATE_DIR = DOCS_DIR / "state"
STATE_PATH = STATE_DIR / "abcd_campaigns.json"

EMAIL_BODY_PATH = Path("email_body.txt")

# -------- XLSX parsing (CAIXA download format) --------

def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _parse_brl(v: object) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace("R$", "").replace(".", "").replace(",", ".").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else None

def _default_cost15_by_date(data_iso: Optional[str]) -> float:
    return 3.50 if (data_iso and data_iso >= "2025-07-10") else 1.00

def _find_header_row(ws) -> Tuple[int, Dict[str, int]]:
    """
    Finds a header row that contains 'Concurso' and 15 ball columns.
    Returns (row_index_1based, col_map) where col_map keys:
      concurso, data, b1..b15
    """
    for r in range(1, 51):
        row = [ws.cell(row=r, column=c).value for c in range(1, 60)]
        norm = [_norm_header(str(v)) if v is not None else "" for v in row]

        try:
            c_conc = norm.index("concurso") + 1
        except ValueError:
            continue

        c_date = None
        for cand in ("data sorteio", "data"):
            if cand in norm:
                c_date = norm.index(cand) + 1
                break

        ball_cols: Dict[str, int] = {}
        for i in range(1, 16):
            # real file uses "Bola1"..."Bola15" (no space) -> normalized becomes "bola1"
            patterns = [
                f"bola{i}",
                f"bola {i}",
                f"bolas {i}",
                f"dezena{i}",
                f"dezena {i}",
                f"{i}ª dezena",
                f"b{i}",
                f"b {i}",
            ]
            found = None
            for p in patterns:
                if p in norm:
                    found = norm.index(p) + 1
                    break
            if found is None:
                ball_cols = {}
                break
            ball_cols[f"b{i}"] = found

        if ball_cols:
            col_map = {"concurso": c_conc}
            if c_date:
                col_map["data"] = c_date
            col_map.update(ball_cols)
            for h in (11, 12, 13, 14, 15):
                key = f"rateio {h} acertos"
                if key in norm:
                    col_map[f"rateio_{h}"] = norm.index(key) + 1
            return r, col_map

    raise RuntimeError("Não consegui localizar o header do XLSX (Concurso/Bola1..Bola15).")

@dataclass
class Draw:
    concurso: int
    data: Optional[str]
    nums: Set[int]
    payouts: Dict[int, float]
    cost15: float

def load_draws_from_xlsx(xlsx_path: str, sheet_name: str = "LOTOFÁCIL") -> List[Draw]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    header_row, cols = _find_header_row(ws)

    draws: List[Draw] = []
    r = header_row + 1
    while True:
        v_conc = ws.cell(row=r, column=cols["concurso"]).value
        if v_conc is None:
            break
        try:
            concurso = int(v_conc)
        except Exception:
            r += 1
            continue

        dval = ws.cell(row=r, column=cols.get("data", cols["concurso"] + 1)).value if "data" in cols else None
        data_str = None
        if isinstance(dval, datetime):
            data_str = dval.date().isoformat()
        elif isinstance(dval, date):
            data_str = dval.isoformat()
        elif isinstance(dval, str) and dval.strip():
            s = dval.strip()
            m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", s)
            data_str = f"{m.group(3)}-{m.group(2)}-{m.group(1)}" if m else s

        nums: Set[int] = set()
        ok = True
        for i in range(1, 16):
            v = ws.cell(row=r, column=cols[f"b{i}"]).value
            if v is None:
                ok = False
                break
            try:
                nums.add(int(v))
            except Exception:
                ok = False
                break

        payouts: Dict[int, float] = {}
        for h in (11, 12, 13, 14, 15):
            c = cols.get(f"rateio_{h}")
            if c:
                pv = _parse_brl(ws.cell(row=r, column=c).value)
                if pv is not None:
                    payouts[h] = float(pv)

        cost15 = _default_cost15_by_date(data_str)
        if payouts.get(11):
            cost15 = round(float(payouts[11]) / 2.0, 2)

        if ok and len(nums) == 15:
            draws.append(Draw(concurso=concurso, data=data_str, nums=nums, payouts=payouts, cost15=cost15))

        r += 1

    draws.sort(key=lambda d: d.concurso)
    return draws

# -------- campaigns state --------

def load_state() -> Dict:
    if not STATE_PATH.exists():
        return {"version": 1, "updated_at": None, "campaigns": []}
    return json.loads(STATE_PATH.read_text(encoding="utf-8"))

def save_state(state: Dict) -> None:
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    state["updated_at"] = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    STATE_PATH.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")

def campaign_key(start_concurso: int, created_on: str) -> str:
    return f"c_{start_concurso}_{created_on.replace('-','')}"

def ensure_snapshot_dirs(ymd: str) -> Path:
    yyyy, mm, _ = ymd.split("-")
    outdir = RESULTS_DIR / yyyy / mm
    outdir.mkdir(parents=True, exist_ok=True)
    return outdir

def write_daily_snapshot(sig: Dict, ymd: str) -> Path:
    outdir = ensure_snapshot_dirs(ymd)
    outpath = outdir / f"{ymd}.json"
    outpath.write_text(json.dumps(sig, ensure_ascii=False, indent=2), encoding="utf-8")
    return outpath

def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser()
    ap.add_argument("--teimosinha", type=int, default=37)
    ap.add_argument("--min_hits_stop", type=int, default=14)
    ap.add_argument("--gate_percentis", type=str, default="25,75")
    ap.add_argument("--script", type=str, default=SCRIPT)
    ap.add_argument("--xlsx", type=str, default="", help="Opcional: path fixo do XLSX (senão usa auto do script).")
    ap.add_argument("--force_email", action="store_true", help="Força gerar email_body.txt mesmo se gate_pass=False e sem campanhas.")
    return ap.parse_args()

def run_daily_signal(script: str, teimosinha: int, min_hits_stop: int, gate_percentis: str) -> Dict:
    cmd = [
        "python", script,
        "--abcd_daily_signal",
        "--abcd_teimosinha_n", str(teimosinha),
        "--abcd_min_hits", str(min_hits_stop),
        "--abcd_gate_percentis", gate_percentis,
        "--abcd_daily_json", "abcd_signal.json",
    ]
    subprocess.check_call(cmd)
    return json.loads(Path("abcd_signal.json").read_text(encoding="utf-8"))

def find_latest_xlsx() -> Optional[str]:
    files = sorted(Path(".").glob("resultados_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return str(files[0]) if files else None

def parse_game_nums(s: str) -> Set[int]:
    nums: Set[int] = set()
    for tok in re.split(r"[,\s]+", (s or "").strip()):
        if not tok:
            continue
        try:
            nums.add(int(tok))
        except Exception:
            pass
    return nums

def compute_hits(game: Set[int], draw_nums: Set[int]) -> int:
    return len(game.intersection(draw_nums))

def check_campaign_against_draw(camp: Dict, draw: Draw) -> Dict:
    jogos = camp.get("jogos", {}) or {}
    parsed_games: List[Tuple[str, Set[int]]] = [(str(k), parse_game_nums(str(v))) for k, v in jogos.items()]

    best_hits = -1
    best_key = None
    best_payout = 0.0
    per_game = []
    paying_events = []
    cost_per_game = float(draw.cost15 or 0.0)

    for k, gnums in parsed_games:
        h = compute_hits(gnums, draw.nums)
        payout = float((draw.payouts or {}).get(h, 0.0)) if h >= 11 else 0.0
        profit = payout - cost_per_game
        rec = {
            "game": k,
            "hits": h,
            "payout": round(payout, 2),
            "cost": round(cost_per_game, 2),
            "profit": round(profit, 2),
        }
        per_game.append(rec)

        if h >= 11:
            paying_events.append(
                {
                    "game": k,
                    "hits": h,
                    "payout": round(payout, 2),
                    "cost": round(cost_per_game, 2),
                    "profit": round(profit, 2),
                }
            )

        if h > best_hits or (h == best_hits and payout > best_payout):
            best_hits = h
            best_key = k
            best_payout = payout

    total_cost = round(cost_per_game * len(parsed_games), 2)
    total_payout = round(sum(float(x.get("payout", 0.0)) for x in per_game), 2)
    total_profit = round(total_payout - total_cost, 2)

    chk = {
        "concurso": draw.concurso,
        "data": draw.data,
        "offset": _camp_offset(camp, draw.concurso),
        "best_hits": best_hits,
        "best_game": best_key,
        "best_payout": round(best_payout, 2),
        "per_game": per_game,
        "paying_events": paying_events,
        # canonical runner fields (kept for backward compatibility)
        "total_cost": total_cost,
        "total_payout": total_payout,
        "total_profit": total_profit,
        # aliases expected by HTML / enriched schema
        "cost_total": total_cost,
        "payout_total": total_payout,
        "profit_total": total_profit,
    }
    return chk

def already_checked(camp: Dict, concurso: int) -> bool:
    return any((chk.get("concurso") == concurso) for chk in (camp.get("checks", []) or []))

def within_offset_window(camp: Dict, concurso: int) -> bool:
    start = int(camp["target_start_concurso"])
    n = int(camp["teimosinha_n"])
    end = start + n - 1
    return start <= concurso <= end

def checks_done_in_window(camp: Dict) -> int:
    return len(camp.get("checks", []) or [])


def _camp_offset(camp: Dict, concurso: int) -> int:
    return int(concurso) - int(camp.get("target_start_concurso", 0)) + 1

def _campaign_totals(camp: Dict) -> Dict[str, float]:
    checks = camp.get("checks", []) or []
    custo = round(sum(float(ch.get("total_cost", 0.0) or 0.0) for ch in checks), 2)
    payout = round(sum(float(ch.get("total_payout", 0.0) or 0.0) for ch in checks), 2)
    return {"custo_total": custo, "payout_total": payout, "profit_total": round(payout - custo, 2)}

def _refresh_campaign_summary(camp: Dict) -> None:
    checks = camp.get("checks", []) or []
    totals = _campaign_totals(camp)

    pay_events = []
    best_hits = -1
    best_game = None
    best_concurso = None

    max_offset = 0
    for ch in checks:
        conc = ch.get("concurso")
        if conc is not None:
            try:
                off = _camp_offset(camp, int(conc))
                ch["offset"] = ch.get("offset", off)
                max_offset = max(max_offset, off)
            except Exception:
                pass

        bh = int(ch.get("best_hits", -1))
        if bh > best_hits:
            best_hits = bh
            best_game = ch.get("best_game")
            best_concurso = ch.get("concurso")

        # propagate paying events (>=11) with campaign-level context
        for ev in (ch.get("paying_events", []) or []):
            pay_events.append(
                {
                    "concurso": ch.get("concurso"),
                    "data": ch.get("data"),
                    "offset": ch.get("offset") or (_camp_offset(camp, int(ch.get("concurso")))),
                    "game": ev.get("game"),
                    "hits": ev.get("hits"),
                    "payout": ev.get("payout"),
                    "cost": ev.get("cost"),  # cost per game on that concurso
                    "profit": ev.get("profit"),
                }
            )

        # ensure check-level aliases exist for HTML/enriched schema
        if "cost_total" not in ch and "total_cost" in ch:
            ch["cost_total"] = ch.get("total_cost")
        if "payout_total" not in ch and "total_payout" in ch:
            ch["payout_total"] = ch.get("total_payout")
        if "profit_total" not in ch and "total_profit" in ch:
            ch["profit_total"] = ch.get("total_profit")

    camp["paying_hits_ge11"] = pay_events

    # offset_atual is the last checked offset inside the teimosinha window
    camp["offset_atual"] = max_offset
    remaining = max(0, int(camp.get("teimosinha_n", 0)) - max_offset)

    camp["summary"] = {
        "checks_count": len(checks),
        "remaining": remaining,
        "best_hits_ate_agora": (None if best_hits < 0 else best_hits),
        "best_game_ate_agora": best_game,
        "best_concurso_ate_agora": best_concurso,
        "eventos_pagantes_ge11": len(pay_events),
        **totals,
    }

def _campaign_needs_backfill(camp: Dict) -> bool:
    # Campaign-level fields
    if not isinstance(camp.get("summary"), dict):
        return True
    if "paying_hits_ge11" not in camp:
        return True

    # Check-level fields (older runs may have missing paying_events / totals)
    for ch in (camp.get("checks", []) or []):
        if "paying_events" not in ch:
            return True
        if "total_cost" not in ch or "total_payout" not in ch or "total_profit" not in ch:
            return True
    return False


def _infer_won_concurso_from_checks(camp: Dict, checks: List[Dict]) -> Optional[int]:
    try:
        stop = int(camp.get("min_hits_stop", 14))
    except Exception:
        stop = 14
    for ch in checks:
        try:
            if int(ch.get("best_hits", -1)) >= stop:
                return int(ch.get("concurso"))
        except Exception:
            continue
    return None


def backfill_campaign_history(camp: Dict, draw_by_concurso: Dict[int, Draw], latest_concurso: int) -> bool:
    '''
    Rebuilds checks/paying events/summary for an existing campaign from target_start_concurso up to:
      - latest_concurso (if active), capped by window end
      - window end (if expired)
      - won.when_concurso (if won), or first >= min_hits_stop found during rebuild
    Preserves campaign.status, but enriches missing fields and normalizes check schemas.
    Returns True if campaign was modified.
    '''
    if not _campaign_needs_backfill(camp):
        # still normalize summary/aliases every run
        _refresh_campaign_summary(camp)
        return False

    status = camp.get("status", "active")
    start = int(camp.get("target_start_concurso", 0))
    n = int(camp.get("teimosinha_n", 0))
    end = int(camp.get("target_end_concurso", start + n - 1))

    # determine eval_end
    eval_end = min(latest_concurso, end)
    if status == "expired":
        eval_end = min(latest_concurso, end)
    elif status == "won":
        when = None
        won = camp.get("won", {}) or {}
        try:
            if won.get("when_concurso") is not None:
                when = int(won.get("when_concurso"))
        except Exception:
            when = None
        if when is not None:
            eval_end = min(when, end, latest_concurso)

    # rebuild sequentially
    new_checks: List[Dict] = []
    for conc in range(start, eval_end + 1):
        d = draw_by_concurso.get(conc)
        if not d:
            continue  # XLSX may not have this concurso yet
        chk = check_campaign_against_draw(camp, d)
        new_checks.append(chk)

        # if campaign is won but missing when_concurso, stop at first win
        if status == "won":
            try:
                if int(chk.get("best_hits", -1)) >= int(camp.get("min_hits_stop", 14)):
                    eval_end = conc
                    break
            except Exception:
                pass

    camp["checks"] = new_checks

    # Rebuild last_check
    if new_checks:
        last = new_checks[-1]
        camp["last_check"] = {
            "concurso": last.get("concurso"),
            "data": last.get("data"),
            "best_hits": last.get("best_hits"),
            "best_game": last.get("best_game"),
            "total_profit": last.get("total_profit"),
        }

    # If status=won and won block incomplete, fill it from rebuilt checks
    if status == "won":
        when = (camp.get("won", {}) or {}).get("when_concurso")
        if when is None:
            when2 = _infer_won_concurso_from_checks(camp, new_checks)
            if when2 is not None:
                # pick the winning check
                win_chk = next((c for c in new_checks if c.get("concurso") == when2), None)
                camp["won"] = {
                    "when_concurso": when2,
                    "best_hits": win_chk.get("best_hits") if win_chk else None,
                    "best_game": win_chk.get("best_game") if win_chk else None,
                    "offset": _camp_offset(camp, when2),
                    "best_payout": win_chk.get("best_payout") if win_chk else None,
                }

    _refresh_campaign_summary(camp)
    return True

def _fmt_games_block(jogos: Dict) -> List[str]:
    out = []
    for k, v in (jogos or {}).items():
        out.append(f"  {k}: {v}")
    return out

def build_email_digest(
    sig: Dict,
    run_ymd: str,
    latest_draw: Draw,
    opened: List[Dict],
    updates: List[Dict],
    won: List[Dict],
    expired: List[Dict],
    active: List[Dict],
) -> str:
    gate = sig.get("gate", {}) or {}

    lines: List[str] = []
    lines.append("Lotofácil ABCD — Resumo Diário")
    lines.append("")
    lines.append(f"Rodado em (Dublin 04:30+): {run_ymd}")
    lines.append(f"Último concurso disponível: {latest_draw.concurso} | Data: {latest_draw.data}")
    lines.append("")
    lines.append(f"gate_pass (hoje): {sig.get('gate_pass')}")
    lines.append(f"Percentis: {gate.get('percentis')} | Faixa: {gate.get('faixa')} | Gap atual: {gate.get('gap_atual')}")
    lines.append("")

    # always show today's suggested games (from daily JSON)
    lines.append("=== JOGOS DO SINAL DE HOJE (JSON) ===")
    jogos_hoje = sig.get("jogos", {}) or {}
    if jogos_hoje:
        lines.extend(_fmt_games_block(jogos_hoje))
    else:
        lines.append("  (nenhum jogo encontrado no JSON)")
    lines.append("")

    if opened:
        lines.append("=== NOVAS CAMPANHAS ABERTAS HOJE ===")
        for c in opened:
            lines.append(f"- {c['id']} | start={c['start_concurso']} -> alvo_início={c['target_start_concurso']} | teimosinha={c['teimosinha_n']} | stop_hits={c['min_hits_stop']}")
            lines.append("  Jogos:")
            lines.extend(_fmt_games_block(c.get("jogos", {}) or {}))
        lines.append("")

    if won:
        lines.append("=== CAMPANHAS ENCERRADAS (GANHOU >= stop_hits) ===")
        for c in won:
            w = c.get("won", {}) or {}
            tot = _campaign_totals(c)
            lines.append(f"- {c['id']} | start={c['start_concurso']} | ganhou no concurso {w.get('when_concurso')} com {w.get('best_hits')} hits ({w.get('best_game')}) | custo={tot['custo_total']:.2f} payout={tot['payout_total']:.2f} profit={tot['profit_total']:.2f}")
        lines.append("")

    if expired:
        lines.append("=== CAMPANHAS ENCERRADAS (EXPIRADAS) ===")
        for c in expired:
            tot = _campaign_totals(c)
            lines.append(f"- {c['id']} | start={c['start_concurso']} | expirou após {c.get('teimosinha_n')} concursos | custo={tot['custo_total']:.2f} payout={tot['payout_total']:.2f} profit={tot['profit_total']:.2f}")
        lines.append("")

    if updates:
        lines.append("=== CHECKS DE HOJE (por campanha) ===")
        for u in updates:
            chk = u["check"]
            lines.append(f"- {u['id']} | concurso {chk['concurso']} | best_hits={chk['best_hits']} | best_game={chk['best_game']} | dia profit={float(chk.get('total_profit',0)):.2f}")
            for ev in (chk.get("paying_events", []) or []):
                lines.append(f"    pagou: {ev.get('game')} | {ev.get('hits')} hits | payout={float(ev.get('payout',0)):.2f} | profit={float(ev.get('profit',0)):.2f}")
        lines.append("")

    if active:
        lines.append("=== CAMPANHAS ATIVAS (LEMBRETE) ===")
        for c in active:
            done = checks_done_in_window(c)
            total = int(c["teimosinha_n"])
            rem = max(0, total - done)
            last_chk = (c.get("checks", []) or [])[-1] if (c.get("checks") or []) else None
            tot = _campaign_totals(c)
            lines.append(f"- {c['id']} | start={c['start_concurso']} -> alvo={c['target_start_concurso']} | surpresinha {done}/{total} | remaining={rem} | acum profit={tot['profit_total']:.2f}")
            if last_chk:
                lines.append(f"  último: concurso {last_chk.get('concurso')} | best_hits={last_chk.get('best_hits')} | best_game={last_chk.get('best_game')}")
            lines.append("  Jogos:")
            lines.extend(_fmt_games_block(c.get("jogos", {}) or {}))
        lines.append("")

    return "\n".join(lines)

def main() -> int:
    args = parse_args()

    # Ensure dirs exist (avoids 404 in GH Pages / html reads)
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    if not STATE_PATH.exists():
        save_state({"version": 1, "updated_at": None, "campaigns": []})

    # Delegate signal+gate to your validated script
    sig = run_daily_signal(args.script, args.teimosinha, args.min_hits_stop, args.gate_percentis)

    # Snapshot naming: use "today in Dublin run" date (UTC is fine for file partitioning),
    # but we keep the script's last_data for context.
    run_ymd = datetime.now(ZoneInfo("Europe/Dublin")).date().isoformat()
    # Find XLSX saved by your script auto-download flow
    xlsx = args.xlsx.strip() or find_latest_xlsx()
    if not xlsx:
        raise RuntimeError("Não encontrei resultados_*.xlsx no workspace após rodar o daily_signal.")

    draws = load_draws_from_xlsx(xlsx)
    if not draws:
        raise RuntimeError("Não consegui carregar concursos do XLSX.")
    latest = draws[-1]

    sig.setdefault("config", {})
    sig["config"]["teimosinha_n"] = int(args.teimosinha)
    sig["config"]["min_hits_stop"] = int(args.min_hits_stop)
    sig["config"]["gate_percentis"] = args.gate_percentis
    try:
        sig["gap_dentro_teimosinha"] = int((sig.get("gate") or {}).get("gap_atual", -1)) <= int(args.teimosinha)
    except Exception:
        sig["gap_dentro_teimosinha"] = None

    state = load_state()
    campaigns: List[Dict] = state.get("campaigns", []) or []

    # Map draws by concurso for fast backfill/checks
    draw_by_concurso: Dict[int, Draw] = {d.concurso: d for d in draws}

    # Backfill/enrich existing campaigns that were created by older runner versions
    backfilled: List[str] = []
    for c in campaigns:
        try:
            if backfill_campaign_history(c, draw_by_concurso, latest.concurso):
                backfilled.append(str(c.get("id")))
        except Exception:
            # Never break daily run due to a single campaign
            pass

    opened: List[Dict] = []
    won: List[Dict] = []
    expired: List[Dict] = []
    updates: List[Dict] = []

    # A) Open new campaign if gate_pass=True today
    if bool(sig.get("gate_pass")):
        start_conc = int(sig.get("last_concurso"))
        target_start = start_conc + 1
        created_on = run_ymd
        cid = campaign_key(start_conc, created_on)

        # dedupe by start_concurso OR target_start_concurso
        exists = any(
            int(c.get("start_concurso", -1)) == start_conc or int(c.get("target_start_concurso", -1)) == target_start
            for c in campaigns
        )

        if not exists:
            new_c = {
                "id": cid,
                "status": "active",
                "created_on": created_on,
                "start_concurso": start_conc,
                "target_start_concurso": target_start,
                "target_end_concurso": target_start + int(args.teimosinha) - 1,
                "teimosinha_n": int(args.teimosinha),
                "min_hits_stop": int(args.min_hits_stop),
                "jogos": sig.get("jogos", {}) or {},
                "checks": [],
                "won": {"when_concurso": None, "best_hits": None, "best_game": None},
            }
            _refresh_campaign_summary(new_c)
            campaigns.append(new_c)
            opened.append(new_c)

    # B) Evaluate all active campaigns against latest concurso (offset window)
    for c in campaigns:
        if c.get("status") != "active":
            continue

        if not within_offset_window(c, latest.concurso):
            start = int(c["target_start_concurso"])
            end = start + int(c["teimosinha_n"]) - 1
            if latest.concurso > end:
                c["status"] = "expired"
                expired.append(c)
            continue

        if already_checked(c, latest.concurso):
            continue

        chk = check_campaign_against_draw(c, latest)
        c.setdefault("checks", []).append(chk)
        c["last_check"] = {"concurso": chk.get("concurso"), "data": chk.get("data"), "best_hits": chk.get("best_hits"), "best_game": chk.get("best_game"), "total_profit": chk.get("total_profit")}
        _refresh_campaign_summary(c)
        updates.append({
            "campaign_id": c["id"],
            "id": c["id"],  # compat
            "concurso": chk.get("concurso"),
            "data": chk.get("data"),
            "offset": chk.get("offset"),
            "best_hits": chk.get("best_hits"),
            "best_game": chk.get("best_game"),
            "best_payout": chk.get("best_payout"),
            "total_cost": chk.get("total_cost"),
            "total_payout": chk.get("total_payout"),
            "total_profit": chk.get("total_profit"),
            "check": chk,  # keep full detail for debug
        })

        if int(chk["best_hits"]) >= int(c["min_hits_stop"]):
            c["status"] = "won"
            c["won"] = {
                "when_concurso": chk["concurso"],
                "best_hits": chk["best_hits"],
                "best_game": chk["best_game"],
                "offset": _camp_offset(c, int(chk["concurso"])),
                "best_payout": chk.get("best_payout", 0.0),
            }
            won.append(c)
        else:
            if checks_done_in_window(c) >= int(c["teimosinha_n"]):
                c["status"] = "expired"
                expired.append(c)

    for c in campaigns:
        _refresh_campaign_summary(c)

    active = [c for c in campaigns if c.get("status") == "active"]

    sig["campaigns_resume"] = {
        "active": len(active),
        "won": len([c for c in campaigns if c.get("status") == "won"]),
        "expired": len([c for c in campaigns if c.get("status") == "expired"]),
        "active_ids": [c.get("id") for c in active],
    }
    if updates:
        sig["campaign_updates_today"] = updates
    if opened:
        sig["opened_campaigns_today"] = [
            {"id": c.get("id"), "start_concurso": c.get("start_concurso"), "target_start_concurso": c.get("target_start_concurso"), "target_end_concurso": c.get("target_end_concurso")}
            for c in opened
        ]

    snap_path = write_daily_snapshot(sig, run_ymd)
    print(f"Snapshot: {snap_path}")

    state["campaigns"] = campaigns
    save_state(state)

    # Email policy:
    # - If gate_pass True -> always email (campaign opened)
    # - If any active campaigns -> daily reminder email
    # - If won/expired today -> email
    should_email = bool(opened or active or won or expired or args.force_email)

    if should_email:
        body = build_email_digest(
            sig=sig,
            run_ymd=run_ymd,
            latest_draw=latest,
            opened=opened,
            updates=updates,
            won=won,
            expired=expired,
            active=active,
        )
        EMAIL_BODY_PATH.write_text(body, encoding="utf-8")
        print("[EMAIL] will_send=true")
    else:
        print("[EMAIL] will_send=false")

    Path("runner_out.json").write_text(
        json.dumps(
            {
                "run_ymd": run_ymd,
                "last_concurso_json": sig.get("last_concurso"),
                "latest_concurso_xlsx": latest.concurso,
                "gate_pass": bool(sig.get("gate_pass")),
                "email": should_email,
                "opened": len(opened),
                "active": len(active),
                "won": len(won),
                "expired": len(expired),
            },
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
